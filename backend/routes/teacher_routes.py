# backend/routes/teacher_routes.py

from flask import Blueprint, request, jsonify, g

from app import db
from typing import Any, Dict, cast
import math
from models import (
    Student,
    Subject,
    Mark,
    TeacherSubjectAllocation
)
from models import Result, Teacher
from services.result_service import generate_results_for_division

from schemas import EnterMarkSchema, UpdateMarkSchema
from auth import token_required, hash_password, verify_password
from config import GRACE_MAX
import io
from typing import TYPE_CHECKING

# Import openpyxl at runtime if available; expose name for runtime checks.
if TYPE_CHECKING:
    import openpyxl  # type: ignore
else:
    try:
        import openpyxl
    except Exception:
        openpyxl = None  # type: ignore

teacher_bp = Blueprint("teacher", __name__, url_prefix="/teacher")

enter_mark_schema = EnterMarkSchema()
update_mark_schema = UpdateMarkSchema()


# ======================================================
# Helper: check teacher allocation
# ======================================================
def _check_teacher_allocation(teacher_id, subject_id, division):
    # Exact subject+division allocation
    alloc = TeacherSubjectAllocation.query.filter_by(
        teacher_id=teacher_id,
        subject_id=subject_id,
        division=division
    ).first()
    if alloc:
        return alloc

    # Fallback: if the subject is a CORE/main subject, allow any teacher
    # who has an allocation in the same division (so teachers assigned to
    # the division can enter core subject marks without explicit per-subject
    # allocation). This keeps optional-subject checks strict.
    try:
        subject = Subject.query.get(subject_id)
        # Only apply the CORE fallback for numeric-mark subjects. Grade-only
        # subjects (PE/EVS) must be handled by the exact allocation above.
        if subject and subject.subject_eval_type == "MARKS" and (subject.subject_type or "").upper() == "CORE":
            any_alloc = TeacherSubjectAllocation.query.filter_by(
                teacher_id=teacher_id,
                division=division
            ).first()
            if any_alloc:
                return any_alloc
    except Exception:
        pass

    return None


# ======================================================
# Helpers: determine if all marks for a subject/division exist
# ======================================================
def _eligible_student_count_for_subject(subject_id, division):
    subject = Subject.query.get(subject_id)
    if not subject:
        return 0

    query = Student.query.filter_by(division=division, batch_id=g.active_batch)
    if subject.subject_code in ("HINDI", "IT"):
        query = query.filter(Student.optional_subject == subject.subject_code)
    if subject.subject_code in ("MATHS", "SP"):
        query = query.filter(Student.optional_subject_2 == subject.subject_code)

    return query.count()


def _marks_count_for_subject(subject_id, division):
    return Mark.query.filter_by(subject_id=subject_id, division=division, batch_id=g.active_batch).count()


def _are_all_marks_submitted(subject_id, division):
    eligible = _eligible_student_count_for_subject(subject_id, division)
    if eligible == 0:
        return False
    return _marks_count_for_subject(subject_id, division) >= eligible


# ======================================================
# 1️⃣ Get students for a subject & division
# ======================================================
@teacher_bp.route("/students", methods=["GET"])
@token_required
def get_students(user_id=None, user_type=None):
    """
    Returns students visible to the logged-in teacher
    based on subject + division allocation
    """
    subject_code = request.args.get("subject_code")
    division = request.args.get("division")

    if not subject_code or not division:
        return {"error": "subject_code and division are required"}, 400

    subject = Subject.query.filter_by(subject_code=subject_code).first()
    if not subject:
        return {"error": "Invalid subject"}, 404

    allocation = _check_teacher_allocation(
        user_id, subject.subject_id, division
    )
    if not allocation:
        return {"error": "Not authorized for this subject/division"}, 403

    query = Student.query.filter_by(division=division, batch_id=g.active_batch)

    # Optional subject filtering
    if subject.subject_code in ("HINDI", "IT"):
        query = query.filter(
            Student.optional_subject == subject.subject_code
        )

    if subject.subject_code in ("MATHS", "SP"):
        query = query.filter(
            Student.optional_subject_2 == subject.subject_code
        )

    students = query.order_by(Student.roll_no).all()

    return jsonify([
        {
            "roll_no": s.roll_no,
            "name": s.name,
            "division": s.division,
            "optional_subject": s.optional_subject,
            "optional_subject_2": s.optional_subject_2,
        }
        for s in students
    ]), 200


# ======================================================
# 2️⃣ Enter marks (CREATE)
# ======================================================
@teacher_bp.route("/marks", methods=["POST"])
@token_required
def enter_marks(user_id=None, user_type=None):
    """
    Teacher enters marks for a student (first time)
    """
    data = cast(Dict[str, Any], enter_mark_schema.load(request.json or {}))

    subject = Subject.query.get(data.get("subject_id"))
    if not subject:
        return {"error": "Invalid subject"}, 404

    allocation = _check_teacher_allocation(
        user_id, subject.subject_id, data.get("division")
    )
    if not allocation:
        return {"error": "Not authorized"}, 403

    student = Student.query.filter_by(
        roll_no=data.get("roll_no"),
        division=data.get("division")
    ).first()

    if not student:
        return {"error": "Student not found"}, 404

    # Optional subject validation
    if subject.subject_code in ("HINDI", "IT"):
        if student.optional_subject != subject.subject_code:
            return {"error": "Student not enrolled in this optional subject"}, 400

    if subject.subject_code in ("MATHS", "SP"):
        if student.optional_subject_2 != subject.subject_code:
            return {"error": "Student not enrolled in this optional subject"}, 400

    existing = Mark.query.filter_by(
        roll_no=data["roll_no"],
        division=data["division"],
        subject_id=subject.subject_id,
        batch_id=g.active_batch
    ).first()

    if existing:
        return {"error": "Marks already exist. Use update instead."}, 409

    # server-side validation for allowed ranges (schemas already validate ranges,
    # but double-check here and enforce grace max)
    unit1 = float(data.get("unit1", 0))
    unit2 = float(data.get("unit2", 0))
    term = float(data.get("term", 0))
    annual = float(data.get("annual", 0))
    internal = float(data.get("internal", 0)) if "internal" in data else 0.0

    if unit1 < 0 or unit1 > 25 or unit2 < 0 or unit2 > 25:
        return {"error": "Unit1 and Unit2 must be between 0 and 25"}, 400
    if term < 0 or term > 50:
        return {"error": "Terminal marks must be between 0 and 50"}, 400
    if annual < 0 or annual > 100:
        return {"error": "Annual marks must be between 0 and 100"}, 400
    if internal < 0 or internal > GRACE_MAX:
        return {"error": f"Internal must be between 0 and {GRACE_MAX}"}, 400

    # Grace is optional and may be provided at any time; limit enforced above.

    # Calculate totals (include internal in total and averages)
    tot = unit1 + unit2 + term + annual + internal
    sub_avg = math.ceil(tot / 2)  # normalize to 100

    mark = Mark()
    mark.roll_no = data.get("roll_no")
    mark.division = data.get("division")
    mark.subject_id = subject.subject_id
    mark.unit1 = data.get("unit1", 0)
    mark.unit2 = data.get("unit2", 0)
    mark.term = data.get("term", 0)
    mark.annual = data.get("annual", 0)
    mark.tot = tot
    mark.sub_avg = sub_avg
    mark.internal = internal
    mark.entered_by = user_id
    mark.batch_id = g.active_batch

    db.session.add(mark)
    db.session.commit()

    # Trigger result generation/update for this division
    try:
        generate_results_for_division(data.get("division"), g.active_batch)
    except Exception as e:
        print(f"Error generating results: {e}")

    return {"message": "Marks entered successfully"}, 201


# ======================================================
# 3️⃣ Update marks
# ======================================================
@teacher_bp.route("/marks/<int:mark_id>", methods=["PUT"])
@token_required
def update_marks(mark_id, user_id=None, user_type=None):
    """
    Update existing marks
    """
    data = cast(Dict[str, Any], update_mark_schema.load(request.json or {}))

    mark = Mark.query.get(mark_id)
    if not mark:
        return {"error": "Marks not found"}, 404

    subject = Subject.query.get(mark.subject_id)
    if not subject:
        return {"error": "Invalid subject"}, 404

    allocation = _check_teacher_allocation(
        user_id, subject.subject_id, mark.division
    )
    if not allocation:
        return {"error": "Not authorized"}, 403

    # validate ranges again server-side
    unit1 = float(data.get("unit1", 0))
    unit2 = float(data.get("unit2", 0))
    term = float(data.get("term", 0))
    annual = float(data.get("annual", 0))
    internal = float(data.get("internal", 0)) if "internal" in data else getattr(mark, 'internal', 0)

    if unit1 < 0 or unit1 > 25 or unit2 < 0 or unit2 > 25:
        return {"error": "Unit1 and Unit2 must be between 0 and 25"}, 400
    if term < 0 or term > 50:
        return {"error": "Terminal marks must be between 0 and 50"}, 400
    if annual < 0 or annual > 100:
        return {"error": "Annual marks must be between 0 and 100"}, 400
    if internal < 0 or internal > GRACE_MAX:
        return {"error": f"Internal must be between 0 and {GRACE_MAX}"}, 400

    tot = unit1 + unit2 + term + annual + internal
    sub_avg = math.ceil(tot / 2)

    mark.unit1 = data["unit1"]
    mark.unit2 = data["unit2"]
    mark.term = data["term"]
    mark.annual = data["annual"]
    mark.tot = tot
    mark.sub_avg = sub_avg
    # If attempting to change internal marks, only allow when all marks for the
    # subject/division have been submitted.
    if "internal" in data:
        # allow internal to be set/updated at any time; enforce allowed range
        g = data.get("internal", getattr(mark, 'internal', 0))
        if g is None:
            g = 0
        if float(g) < 0 or float(g) > GRACE_MAX:
            return {"error": f"Internal must be between 0 and {GRACE_MAX}"}, 400
        mark.internal = g

    db.session.commit()

    db.session.commit()

    # Trigger result generation/update for this division
    try:
        generate_results_for_division(mark.division, g.active_batch)
    except Exception as e:
        print(f"Error generating results: {e}")

    return {"message": "Marks updated successfully"}, 200


@teacher_bp.route("/marks", methods=["GET"])
@token_required
def list_marks(user_id=None, user_type=None):
    """
    List marks for a given subject_id and division. Returns rows for all students
    (including those without marks) to make frontend table rendering simple.
    Query params: subject_id, division
    """
    subject_id_raw = request.args.get("subject_id")
    division = request.args.get("division")

    if not subject_id_raw or not division:
        return {"error": "subject_id and division are required"}, 400

    try:
        subject_id = int(subject_id_raw)
    except Exception:
        return {"error": "Invalid subject_id"}, 400

    # ensure teacher allocation (use centralized helper which allows
    # CORE subject fallback based on any allocation in the division)
    alloc = _check_teacher_allocation(user_id, subject_id, division)
    if not alloc and user_type != "ADMIN":
        # attempt relaxed fallbacks for CORE subjects: if the teacher has any
        # allocation in the same division or any allocation at all, allow read.
        subj = Subject.query.get(subject_id)
        if subj and (subj.subject_type or "").upper() == "CORE":
            same_div = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=division).first()
            any_alloc = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id).first()
            if same_div or any_alloc:
                alloc = same_div or any_alloc

    if not alloc and user_type != "ADMIN":
        return {"error": "Not authorized for this subject/division"}, 403

    # If the subject is optional, restrict to students who selected it.
    students_query = Student.query.filter_by(division=division, batch_id=g.active_batch)
    # determine subject code for optional filtering
    subj = Subject.query.get(subject_id)
    if subj and subj.subject_code in ("HINDI", "IT"):
        students_query = students_query.filter(Student.optional_subject == subj.subject_code)
    if subj and subj.subject_code in ("MATHS", "SP"):
        students_query = students_query.filter(Student.optional_subject_2 == subj.subject_code)

    students = students_query.order_by(Student.roll_no).all()

    # map existing marks by roll_no
    marks = Mark.query.filter_by(subject_id=subject_id, division=division, batch_id=g.active_batch).all()
    marks_map = {m.roll_no: m for m in marks}

    rows = []
    for s in students:
        m = marks_map.get(s.roll_no)
        rows.append({
            "roll_no": s.roll_no,
            "name": s.name,
            "division": s.division,
            "optional_subject": s.optional_subject,
            "optional_subject_2": s.optional_subject_2,
            "mark": {
                "mark_id": m.mark_id if m else None,
                "unit1": m.unit1 if m else None,
                "unit2": m.unit2 if m else None,
                "term": m.term if m else None,
                "annual": m.annual if m else None,
                "tot": m.tot if m else None,
                "total": m.tot if m else None,
                "sub_avg": m.sub_avg if m else None,
                "average": m.sub_avg if m else None,
                "internal": getattr(m, 'internal', 0) if m else 0,
                # Provide `grace` for compatibility: prefer explicit `grace` column,
                # otherwise fall back to `internal` which teachers edit.
                "grace": (getattr(m, 'grace', None) if m and getattr(m, 'grace', None) is not None else (getattr(m, 'internal', 0) if m else 0)),
            }
        })

    return jsonify(rows), 200


@teacher_bp.route('/divisions', methods=['GET'])
@token_required
def teacher_divisions(user_id=None, user_type=None):
    """Return divisions visible to the teacher.
    If a subject_id (id or code) is provided and the subject is CORE, return all divisions.
    Otherwise return divisions where the teacher has allocations.
    """
    subject_q = request.args.get('subject_id')
    # try resolve subject
    subj = None
    if subject_q:
        try:
            sid = int(subject_q)
            subj = Subject.query.get(sid)
        except Exception:
            try:
                sv = str(subject_q).strip()
                subj = Subject.query.filter((Subject.subject_code == sv) | (Subject.subject_name == sv)).first()
            except Exception:
                subj = None

    if subj and (subj.subject_type or '').upper() == 'CORE':
        # return all divisions (teacher can view core subject marks across divisions)
        divs = [d[0] for d in db.session.query(Student.division).distinct().all()]
        return jsonify(sorted(divs)), 200

    # otherwise, return divisions teacher has allocations for
    allocs = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id).all()
    divs = sorted({a.division for a in allocs})
    return jsonify(divs), 200


@teacher_bp.route("/marks/<int:mark_id>", methods=["DELETE"])
@token_required
def delete_mark(mark_id, user_id=None, user_type=None):
    """
    Delete a mark row. Only the teacher who entered it or an ADMIN can delete.
    """
    mark = Mark.query.get(mark_id)
    if not mark:
        return {"error": "Marks not found"}, 404

    # authorization: teacher who entered or admin
    if user_type != "ADMIN" and mark.entered_by != user_id:
        return {"error": "Not authorized to delete this mark"}, 403
    # (fixed) stray/garbage line removed

    db.session.delete(mark)
    db.session.commit()

    # Note: Teachers update only the `marks` table. Result regeneration is
    # intentionally skipped here to avoid teachers modifying `Result` rows.

    return {"message": "Marks deleted"}, 200





# ======================================================
# 4️⃣ View Complete Table (per-division)
# Returns student list with subject-wise avg, grace and final marks
# ======================================================
@teacher_bp.route("/complete-table", methods=["GET"])
@token_required
def view_complete_table(user_id=None, user_type=None):
    division = request.args.get("division")
    if not division:
        return {"error": "division is required"}, 400

    # Ensure teacher is authorized for this division (has any allocation)
    alloc = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=division).first()
    if not alloc and user_type != "ADMIN":
        return {"error": "Not authorized for this division"}, 403

    # Do not regenerate results from teacher view requests. Admin or background
    # processes should update the `Result` table when needed.

    # fetch students in canonical order and build rows
    students = Student.query.filter_by(division=division, batch_id=g.active_batch).order_by(Student.roll_no).all()
    subjects = {s.subject_id: s.subject_code for s in Subject.query.all()}

    rows = []
    for idx, s in enumerate(students, start=1):
        result = None
        from models import Result
        result = Result.query.filter_by(roll_no=s.roll_no, division=s.division, batch_id=g.active_batch).first()

        # build per-subject entries from Result columns
        subject_entries = []
        total_avg = 0
        total_internal = 0

        # Core Subjects
        for code in ["ENG", "ECO", "BK", "OC"]:
            avg, grace = (None, 0.0)
            if result:
                avg, grace = result.get_subject_data(code)
            
            final = None
            if avg is not None:
                final = (avg or 0) + (grace or 0)
                total_avg += avg or 0
                total_internal += grace or 0

            subject_entries.append({"code": code, "avg": avg, "internal": grace, "final": final})

        # Optional Subjects
        # Iterate over potential optionals, check if student has them
        for code in ["HINDI", "IT", "MATHS", "SP"]:
            include = False
            if code in ("HINDI", "IT") and s.optional_subject == code:
                include = True
            if code in ("MATHS", "SP") and s.optional_subject_2 == code:
                include = True
            
            if include:
                avg, grace = (None, 0.0)
                if result:
                    avg, grace = result.get_subject_data(code)
                
                final = None
                if avg is not None:
                    final = (avg or 0) + (grace or 0)
                    total_avg += avg or 0
                    total_internal += grace or 0

                subject_entries.append({"code": code, "avg": avg, "internal": grace, "final": final})

        final_total = None
        if subject_entries:
            final_total = total_avg + total_internal

        rows.append({
            "seq": idx,
            "roll_no": s.roll_no,
            "name": s.name,
            "subjects": subject_entries,
            "total_avg": round(total_avg, 2),
            "total_internal": round(total_internal, 2),
            "final_total": round(final_total, 2) if final_total is not None else None,
            "percentage": getattr(result, "percentage", None) if result else None
        })

    return jsonify(rows), 200


@teacher_bp.route('/divisions', methods=['GET'])
@token_required
def list_divisions(user_id=None, user_type=None):
    """Return all distinct divisions present in students table."""
    from sqlalchemy import distinct
    divs = [d[0] for d in db.session.query(distinct(Student.division)).all()]
    return jsonify(sorted([d for d in divs if d is not None])), 200


@teacher_bp.route("/marks/from-excel", methods=["POST"])
@token_required
def marks_from_excel(user_id=None, user_type=None):
    """
    Accept an uploaded Excel file (form-data 'file') and extract roll_no and division
    from the first sheet. Match students in DB by exact roll_no+division and return
    matched and missing lists. Optionally accept form fields 'division' and 'subject_id'
    to supply defaults.
    """
    if openpyxl is None:
        return {"error": "Server missing Excel parsing support (openpyxl)"}, 500

    f = request.files.get('file')
    if not f:
        return {"error": "No file uploaded (file)"}, 400

    default_division = request.form.get('division')
    subject_id_form = request.form.get('subject_id')
    # helper: derive subject_id for this teacher+division when not provided
    def derive_subject_id_for_division(div: str):
        # prefer explicit form value
        if subject_id_form:
            try:
                sid = int(subject_id_form)
                return sid
            except Exception:
                return None
        # find allocations for this teacher in the division
        allocs = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=div).all()
        if not allocs:
            return None
        if len(allocs) == 1:
            return allocs[0].subject_id
        # ambiguous if multiple allocations exist; require explicit subject_id in form
        return None

    try:
        data = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(data, data_only=True)
    except Exception as e:
        return {"error": "Failed to read Excel file", "details": str(e)}, 400

    if not wb.sheetnames:
        return {"error": "Excel contains no sheets"}, 400

    # Use the first sheet (do not enforce exact sheet name). Validate by columns instead.
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(x).strip().lower() if x is not None else '' for x in next(rows_iter)]
    except StopIteration:
        return {"error": "Excel sheet is empty"}, 400

    if not any(headers):
        return {"error": "Excel header row is empty"}, 400

    # identify indices
    def idx_of(names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    # Required template columns (case-insensitive) for the teacher upload step
    # At minimum require: Roll and Division (Name optional). Subject is not required in Excel.
    roll_idx = idx_of(['roll_no', 'roll', 'rollno'])
    name_idx = idx_of(['name', 'student_name', 'student'])
    div_idx = idx_of(['division', 'div'])
    # optional subject column
    subj_idx = idx_of(['subject', 'subject_code', 'subject_id'])
    # optional unit columns
    u1_idx = idx_of(['unit1', 'unit 1', 'unit_i'])
    u2_idx = idx_of(['unit2', 'unit 2', 'unit_ii'])
    term_idx = idx_of(['term', 'terminal', 'term i', 'term_i'])
    annual_idx = idx_of(['annual', 'annual marks', 'annual_marks'])
    grace_idx = idx_of(['internal', 'grace'])

    if roll_idx is None or div_idx is None:
        return {"error": "Excel must include columns: Roll and Division (Name optional)"}, 400
    requested = []
    for r in rows_iter:
        if not r or all(c is None for c in r):
            continue
        roll_val = r[roll_idx] if roll_idx is not None and roll_idx < len(r) else None
        if roll_val is None or str(roll_val).strip() == '':
            continue
        roll = str(roll_val).strip()
        division = None
        if div_idx is not None and div_idx < len(r):
            dv = r[div_idx]
            if dv is not None and str(dv).strip() != '':
                division = str(dv).strip()
        if not division and default_division:
            division = default_division

        name_val = r[name_idx] if name_idx is not None and name_idx < len(r) else None
        name_val = str(name_val).strip() if name_val is not None else None

        # optional subject cell
        subject_val = None
        if subj_idx is not None and subj_idx < len(r):
            sv = r[subj_idx]
            if sv is not None and str(sv).strip() != '':
                subject_val = str(sv).strip()

        # optional marks columns
        def val_at_idx(ix):
            return r[ix] if ix is not None and ix < len(r) else None

        unit1_val = val_at_idx(u1_idx)
        unit2_val = val_at_idx(u2_idx)
        term_val = val_at_idx(term_idx)
        annual_val = val_at_idx(annual_idx)
        grace_val = val_at_idx(grace_idx)

        # build requested entry; subject_id will be derived later (prefer form, then subject cell, then allocation)
        requested.append({
            "roll_no": roll,
            "division": division,
            "name": name_val,
            "subject_val": subject_val,
            "unit1": unit1_val,
            "unit2": unit2_val,
            "term": term_val,
            "annual": annual_val,
            "internal": grace_val,
        })

    if not requested:
        return {"error": "No valid rows found in Excel"}, 400

    matched = []
    missing = []
    for item in requested:
        if not item['division']:
            missing.append({"roll_no": item['roll_no'], "division": None, "reason": "division missing"})
            continue
        student = Student.query.filter_by(roll_no=item['roll_no'], division=item['division'], batch_id=g.active_batch).first()
        if not student:
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "student not found"})
            continue

        # resolve subject for this division
        sid = None
        # 1) form-provided subject
        if subject_id_form:
            try:
                sid = int(subject_id_form)
            except Exception:
                sid = None
        # 2) subject cell in Excel
        if sid is None and item.get('subject_val'):
            sv = item.get('subject_val')
            try:
                if str(sv).isdigit():
                    sid = int(sv)
                else:
                    s = Subject.query.filter((Subject.subject_code == sv) | (Subject.subject_name == sv)).first()
                    if s:
                        sid = s.subject_id
            except Exception:
                sid = None
        # 3) derive from allocation for this division
        if sid is None:
            sid = derive_subject_id_for_division(item['division'])

        if not sid:
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "subject not resolved or ambiguous; provide subject_id"})
            continue

        # Ensure teacher is authorized for this subject+division
        alloc = _check_teacher_allocation(user_id, int(sid), item['division'])
        if not alloc and user_type != 'ADMIN':
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "not authorized for this subject/division"})
            continue

        # prepare mark object preferring excel values where present
        def try_float(v):
            try:
                return float(v)
            except Exception:
                return None

        u1 = try_float(item.get('unit1'))
        u2 = try_float(item.get('unit2'))
        t = try_float(item.get('term'))
        a = try_float(item.get('annual'))
        g = try_float(item.get('internal'))

        row = {"roll_no": student.roll_no, "name": student.name, "division": student.division}
        mark = Mark.query.filter_by(subject_id=sid, roll_no=student.roll_no, division=student.division).first()
        # populate fields: prefer existing DB values, but override with uploaded excel where provided
        row['mark'] = {
            "mark_id": mark.mark_id if mark else None,
            "unit1": u1 if u1 is not None else (mark.unit1 if mark else None),
            "unit2": u2 if u2 is not None else (mark.unit2 if mark else None),
            "term": t if t is not None else (mark.term if mark else None),
            "annual": a if a is not None else (mark.annual if mark else None),
            "tot": mark.tot if mark else ( ( (u1 or 0) + (u2 or 0) + (t or 0) + (a or 0) + (g or 0)) if any(x is not None for x in (u1,u2,t,a,g)) else None ),
            "sub_avg": mark.sub_avg if mark else None,
            "internal": g if g is not None else (getattr(mark, 'internal', 0) if mark else 0),
            "grace": (getattr(mark, 'grace', None) if mark and getattr(mark, 'grace', None) is not None else (g if g is not None else (getattr(mark, 'internal', 0) if mark else 0))),
        }
        row['subject_id'] = sid
        matched.append(row)

    return jsonify({"matched": matched, "missing": missing}), 200


@teacher_bp.route('/marks/batch', methods=['POST'])
@token_required
def batch_upsert_marks(user_id=None, user_type=None):
    data = request.get_json() or {}
    entries = data.get('entries')
    if not entries or not isinstance(entries, list):
        return {"error": "entries (array) is required"}, 400

    errors = []
    saved = []
    divisions_to_regen = set()

    for idx, e in enumerate(entries, start=1):
        roll = e.get('roll_no') or e.get('roll')
        division = e.get('division')
        subject_id = e.get('subject_id')
        if not roll or not division or not subject_id:
            errors.append({"index": idx, "error": "roll_no, division and subject_id are required"})
            continue

        student = Student.query.filter_by(roll_no=str(roll), division=division, batch_id=g.active_batch).first()
        if not student:
            errors.append({"index": idx, "roll_no": roll, "division": division, "error": "student not found"})
            continue

        # Resolve subject: accept numeric id or subject code/name strings
        subject = None
        try:
            sid = int(subject_id)
            subject = Subject.query.get(sid)
        except Exception:
            try:
                sv = str(subject_id).strip()
                if sv:
                    s = Subject.query.filter((Subject.subject_code == sv) | (Subject.subject_name == sv)).first()
                    if s:
                        subject = s
                        sid = s.subject_id
            except Exception:
                subject = None

        if not subject:
            errors.append({"index": idx, "roll_no": roll, "division": division, "error": "Invalid subject"})
            continue

        alloc = _check_teacher_allocation(user_id, subject.subject_id, division)
        if not alloc and user_type != 'ADMIN':
            # Relaxed fallback for CORE subjects: allow if teacher has any allocation
            # (or an allocation in the same division). Mirrors the read/list behaviour.
            try:
                if (subject.subject_type or "").upper() == "CORE":
                    same_div = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=division).first()
                    any_alloc = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id).first()
                    if same_div or any_alloc:
                        alloc = same_div or any_alloc
            except Exception:
                pass

        if not alloc and user_type != 'ADMIN':
            errors.append({"index": idx, "roll_no": roll, "division": division, "error": "not authorized for subject/division"})
            continue

        # Ensure student is enrolled in the optional subject when subject is optional
        if subject.subject_code in ("HINDI", "IT") and student.optional_subject != subject.subject_code:
            errors.append({"index": idx, "roll_no": roll, "division": division, "error": "student not enrolled in this optional subject"})
            continue
        if subject.subject_code in ("MATHS", "SP") and student.optional_subject_2 != subject.subject_code:
            errors.append({"index": idx, "roll_no": roll, "division": division, "error": "student not enrolled in this optional subject"})
            continue

        try:
            unit1 = float(e.get('unit1', 0) or 0)
            unit2 = float(e.get('unit2', 0) or 0)
            term = float(e.get('term', 0) or 0)
            annual = float(e.get('annual', 0) or 0)
            internal = float(e.get('internal', 0) or 0)
        except Exception:
            errors.append({"index": idx, "roll_no": roll, "error": "invalid numeric value"})
            continue

        if unit1 < 0 or unit1 > 25 or unit2 < 0 or unit2 > 25 or term < 0 or term > 50 or annual < 0 or annual > 80 or internal < 0 or internal > GRACE_MAX:
            errors.append({"index": idx, "roll_no": roll, "division": division, "error": "one or more marks out of allowed ranges"})
            continue

        existing = Mark.query.filter_by(roll_no=str(roll), division=division, subject_id=subject.subject_id, batch_id=g.active_batch).first()
        tot = unit1 + unit2 + term + annual + internal
        sub_avg = math.ceil(tot / 2)
        if existing:
            existing.unit1 = unit1
            existing.unit2 = unit2
            existing.term = term
            existing.annual = annual
            existing.tot = tot
            existing.sub_avg = sub_avg
            existing.internal = internal
        else:
            m = Mark()
            m.roll_no = str(roll)
            m.division = division
            m.subject_id = subject.subject_id
            m.unit1 = unit1
            m.unit2 = unit2
            m.term = term
            m.annual = annual
            m.tot = tot
            m.sub_avg = sub_avg
            m.internal = internal
            m.entered_by = user_id
            m.batch_id = g.active_batch
            db.session.add(m)

        divisions_to_regen.add(division)
        saved.append({"roll_no": str(roll), "division": division, "subject_id": subject.subject_id})

    # Only return error if ALL entries failed validation
    if errors and not saved:
        return {"error": "Validation failed for all rows", "details": errors}, 400

    try:
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return {"error": "Database commit failed", "details": str(ex)}, 500

    # Results regeneration skipped for teacher batch upserts. Leaving `marks`
    # table writes only; admin or scheduled tasks should regenerate `Result`.

    # Return success even if some rows had errors, as long as at least one was saved
    response = {"message": "Marks saved successfully", "saved": saved}
    if errors:
        response["validation_warnings"] = errors
    return response, 200


@teacher_bp.route('/marks/upload-apply', methods=['POST'])
@token_required
def upload_apply_excel(user_id=None, user_type=None):
    """Accept master Excel, validate strict template and teacher allocation, then apply marks.
    Returns saved and missing lists."""
    if openpyxl is None:
        return {"error": "Server missing Excel parsing support (openpyxl)"}, 500

    f = request.files.get('file')
    if not f:
        return {"error": "No file uploaded (file)"}, 400

    try:
        data = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(data, data_only=True)
    except Exception as e:
        return {"error": "Failed to read Excel file", "details": str(e)}, 400

    # Use first sheet and accept dynamic columns. Require Roll and Division at minimum.
    sheet_name = wb.sheetnames[0] if wb.sheetnames else None
    if not sheet_name:
        return {"error": "Excel contains no sheets"}, 400

    sheet = wb[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(x).strip().lower() if x is not None else '' for x in next(rows_iter)]
    except StopIteration:
        return {"error": "Excel sheet is empty"}, 400

    def find_header(name_variants):
        for v in name_variants:
            if v in headers:
                return headers.index(v)
        return None

    indices = {
        'roll': find_header(['roll', 'roll_no', 'rollno']),
        'name': find_header(['student name', 'name', 'student_name']),
        'subject': find_header(['subject', 'subject_code', 'subject_id']),
        'division': find_header(['division', 'div']),
        'unit1': find_header(['unit1', 'unit 1', 'unit_i']),
        'unit2': find_header(['unit2', 'unit 2', 'unit_ii']),
        'term': find_header(['term', 'terminal', 'term_i']),
        'annual': find_header(['annual', 'annual marks', 'annual_marks']),
        'internal': find_header(['internal', 'grace'])
    }

    if indices['roll'] is None or indices['division'] is None:
        return {"error": "Invalid Excel template. Missing required columns: roll and division"}, 400

    requested = []
    for r in rows_iter:
        if not r or all(c is None for c in r):
            continue
        roll_val = r[indices['roll']] if indices['roll'] is not None and indices['roll'] < len(r) else None
        if roll_val is None or str(roll_val).strip() == '':
            continue
        roll = str(roll_val).strip()
        division = r[indices['division']] if indices['division'] is not None and indices['division'] < len(r) else None
        division = str(division).strip() if division is not None else None

        subj_val = r[indices['subject']] if indices['subject'] is not None and indices['subject'] < len(r) else None
        subject_id = None
        if subj_val is not None and str(subj_val).strip() != '':
            sv = str(subj_val).strip()
            try:
                if sv.isdigit():
                    subject_id = int(sv)
                else:
                    s = Subject.query.filter((Subject.subject_code == sv) | (Subject.subject_name == sv)).first()
                    if s:
                        subject_id = s.subject_id
            except Exception:
                subject_id = None

        def val_at_key(k):
            idx = indices.get(k)
            return r[idx] if idx is not None and idx < len(r) else None

        unit1 = val_at_key('unit1')
        unit2 = val_at_key('unit2')
        term = val_at_key('term')
        annual = val_at_key('annual')
        internal = val_at_key('internal')

        requested.append({
            'roll_no': roll,
            'division': division,
            'subject_id': subject_id,
            'unit1': unit1,
            'unit2': unit2,
            'term': term,
            'annual': annual,
            'internal': internal
        })

    if not requested:
        return {"error": "No valid rows found in Excel"}, 400

    # validate and filter by teacher allocation; derive subject if needed
    to_apply = []
    missing = []
    for item in requested:
        if not item['division']:
            missing.append({"roll_no": item['roll_no'], "division": None, "reason": "division missing"})
            continue
        student = Student.query.filter_by(roll_no=item['roll_no'], division=item['division'], batch_id=g.active_batch).first()
        if not student:
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "student not found"})
            continue

        # resolve subject: prefer form, then subject cell, then derive from allocation
        sid = None
        sid_val = request.form.get('subject_id')
        if sid_val:
            try:
                sid = int(sid_val)
            except Exception:
                sid = None
        if sid is None and item.get('subject_id'):
            try:
                sid = int(item.get('subject_id'))
            except Exception:
                sid = None
        if sid is None:
            # derive via allocations for this teacher+division
            allocs = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=item['division']).all()
            if allocs and len(allocs) == 1:
                sid = allocs[0].subject_id
            else:
                sid = None

        if not sid:
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "subject not resolved or ambiguous; provide subject_id"})
            continue

        alloc = _check_teacher_allocation(user_id, int(sid), item['division'])
        if not alloc and user_type != 'ADMIN':
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "not authorized for this subject/division"})
            continue

        # numeric conversions: treat missing values as 0
        try:
            u1 = float(item['unit1']) if item.get('unit1') not in (None, '') else 0
            u2 = float(item['unit2']) if item.get('unit2') not in (None, '') else 0
            t = float(item['term']) if item.get('term') not in (None, '') else 0
            a = float(item['annual']) if item.get('annual') not in (None, '') else 0
            g = float(item['internal']) if item.get('internal') not in (None, '') else 0
        except Exception:
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "invalid numeric value"})
            continue

        # range checks
        if u1 < 0 or u1 > 25 or u2 < 0 or u2 > 25 or t < 0 or t > 50 or a < 0 or a > 100 or g < 0 or g > GRACE_MAX:
            missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "marks out of allowed ranges"})
            continue

        # Ensure student is enrolled in optional subject if applicable
        subj_obj = Subject.query.get(int(sid))
        if subj_obj:
            if subj_obj.subject_code in ("HINDI", "IT") and student.optional_subject != subj_obj.subject_code:
                missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "student not enrolled in this optional subject"})
                continue
            if subj_obj.subject_code in ("MATHS", "SP") and student.optional_subject_2 != subj_obj.subject_code:
                missing.append({"roll_no": item['roll_no'], "division": item['division'], "reason": "student not enrolled in this optional subject"})
                continue

        to_apply.append({
            'roll_no': item['roll_no'],
            'division': item['division'],
            'subject_id': int(sid),
            'unit1': u1,
            'unit2': u2,
            'term': t,
            'annual': a,
            'internal': g
        })

    if not to_apply:
        return {"error": "No rows authorized/valid to apply", "missing": missing}, 400

    # Apply upserts in transaction
    saved = []
    divisions_to_regen = set()
    try:
        for e in to_apply:
            existing = Mark.query.filter_by(roll_no=str(e['roll_no']), division=e['division'], subject_id=int(e['subject_id']), batch_id=g.active_batch).first()
            tot = e['unit1'] + e['unit2'] + e['term'] + e['annual'] + e.get('internal', 0)
            sub_avg = math.ceil(tot / 2)
            if existing:
                existing.unit1 = e['unit1']
                existing.unit2 = e['unit2']
                existing.term = e['term']
                existing.annual = e['annual']
                existing.tot = tot
                existing.sub_avg = sub_avg
                existing.internal = e.get('internal', 0)
            else:
                m = Mark()
                m.roll_no = str(e['roll_no'])
                m.division = e['division']
                m.subject_id = int(e['subject_id'])
                m.unit1 = e['unit1']
                m.unit2 = e['unit2']
                m.term = e['term']
                m.annual = e['annual']
                m.tot = tot
                m.sub_avg = sub_avg
                m.internal = e.get('internal', 0)
                m.entered_by = user_id
                m.batch_id = g.active_batch
                db.session.add(m)
            divisions_to_regen.add(e['division'])
            saved.append({"roll_no": str(e['roll_no']), "division": e['division'], "subject_id": int(e['subject_id'])})

        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return {"error": "Database commit failed", "details": str(ex)}, 500

    # Results regeneration skipped for excel apply/upload operations initiated by teachers.

    return {"message": "Marks applied successfully", "saved": saved, "missing": missing}, 200


@teacher_bp.route("/students-by-division", methods=["GET"])
@token_required
def students_by_division(user_id=None, user_type=None):
    """
    Return list of students for a division. Teacher must have any allocation for that division.
    Query param: division
    """
    division = request.args.get("division")
    if not division:
        return {"error": "division is required"}, 400

    # ensure teacher has allocation for this division
    alloc = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=division).first()
    if not alloc and user_type != "ADMIN":
        return {"error": "Not authorized for this division"}, 403

    students = Student.query.filter_by(division=division, batch_id=g.active_batch).order_by(Student.roll_no).all()
    return jsonify([{"roll_no": s.roll_no, "name": s.name} for s in students]), 200


@teacher_bp.route("/student-marks", methods=["GET"])
@token_required
def student_marks(user_id=None, user_type=None):
    """
    Return all subjects for a student (including optional ones) and existing marks if any.
    Query params: roll_no, division
    """
    roll_no = request.args.get("roll_no")
    division = request.args.get("division")
    if not roll_no or not division:
        return {"error": "roll_no and division are required"}, 400

    # ensure teacher has allocation for this division
    alloc = TeacherSubjectAllocation.query.filter_by(teacher_id=user_id, division=division).first()
    if not alloc and user_type != "ADMIN":
        return {"error": "Not authorized for this division"}, 403

    student = Student.query.filter_by(roll_no=roll_no, division=division, batch_id=g.active_batch).first()
    if not student:
        return {"error": "Student not found"}, 404

    # all active subjects
    all_subjects = Subject.query.filter_by(active=True).order_by(Subject.subject_code).all()

    # determine which optional subjects the student takes
    include_codes = set()
    for s in all_subjects:
        if s.subject_type == 'CORE':
            include_codes.add(s.subject_code)
    if student.optional_subject:
        include_codes.add(student.optional_subject)
    if student.optional_subject_2:
        include_codes.add(student.optional_subject_2)

    subjects = [s for s in all_subjects if s.subject_code in include_codes]

    # map marks
    marks = Mark.query.filter_by(roll_no=roll_no, division=division).all()
    marks_map = {m.subject_id: m for m in marks}

    rows = []
    for s in subjects:
        m = marks_map.get(s.subject_id)
        rows.append({
            "subject_id": s.subject_id,
            "subject_code": s.subject_code,
            "subject_name": s.subject_name,
                "mark": {
                "mark_id": m.mark_id if m else None,
                "unit1": m.unit1 if m else None,
                "unit2": m.unit2 if m else None,
                "term": m.term if m else None,
                "annual": m.annual if m else None,
                "tot": m.tot if m else None,
                "sub_avg": m.sub_avg if m else None,
                "internal": getattr(m, 'internal', 0) if m else 0,
                "grace": (getattr(m, 'grace', None) if m and getattr(m, 'grace', None) is not None else (getattr(m, 'internal', 0) if m else 0)),
            }
        })

    return jsonify({"roll_no": roll_no, "name": student.name, "division": division, "subjects": rows}), 200


# =====================================================
# Grade-only subject endpoints (PE / EVS)
# =====================================================


@teacher_bp.route("/grades", methods=["GET"])
@token_required
def list_grades(user_id=None, user_type=None):
    """
    List students for a grade-only subject (PE / EVS) in a division and their current grade.
    Query params: subject_code (PE or EVS), division
    """
    subject_code = request.args.get("subject_code")
    division = request.args.get("division")

    if not subject_code or not division:
        return {"error": "subject_code and division are required"}, 400

    subj = Subject.query.filter_by(subject_code=subject_code).first()
    if not subj:
        return {"error": "Invalid subject_code"}, 404

    if subj.subject_eval_type != "GRADE":
        return {"error": "Subject is not grade-only"}, 400

    # Require explicit allocation for grade subjects
    alloc = TeacherSubjectAllocation.query.filter_by(
        teacher_id=user_id, subject_id=subj.subject_id, division=division
    ).first()
    if not alloc and user_type != "ADMIN":
        return {"error": "Not authorized for this subject/division"}, 403

    students = Student.query.filter_by(division=division, batch_id=g.active_batch).order_by(Student.roll_no).all()

    rows = []
    for s in students:
        res = Result.query.filter_by(batch_id=g.active_batch, roll_no=s.roll_no, division=division).first()
        grade = None
        if res:
            grade = getattr(res, f"{subject_code.lower()}_grade", None)
        rows.append({
            "roll_no": s.roll_no,
            "name": s.name,
            "division": s.division,
            "grade": grade
        })

    return jsonify(rows), 200


@teacher_bp.route("/grades", methods=["POST"])
@token_required
def update_grades(user_id=None, user_type=None):
    """
    Update grades for a grade-only subject. Accepts JSON: { subject_code: 'PE', entries: [{roll_no, division, grade}, ...] }
    Only allocated teacher (exact subject+division) or ADMIN may update.
    This updates the `Result` table (creates if missing) and does NOT touch `marks`.
    """
    data = request.get_json() or {}
    subject_code = data.get("subject_code")
    entries = data.get("entries")

    if not subject_code or not entries or not isinstance(entries, list):
        return {"error": "subject_code and entries (array) are required"}, 400

    subj = Subject.query.filter_by(subject_code=subject_code).first()
    if not subj:
        return {"error": "Invalid subject_code"}, 404
    if subj.subject_eval_type != "GRADE":
        return {"error": "Subject is not grade-only"}, 400

    saved = []
    errors = []
    try:
        for idx, e in enumerate(entries, start=1):
            roll = e.get("roll_no")
            division = e.get("division")
            grade = e.get("grade")
            if not roll or not division:
                errors.append({"index": idx, "error": "roll_no and division required"})
                continue

            # authorization: exact allocation required
            alloc = TeacherSubjectAllocation.query.filter_by(
                teacher_id=user_id, subject_id=subj.subject_id, division=division
            ).first()
            if not alloc and user_type != "ADMIN":
                errors.append({"index": idx, "roll_no": roll, "division": division, "error": "not authorized for this subject/division"})
                continue

            student = Student.query.filter_by(roll_no=roll, division=division, batch_id=g.active_batch).first()
            if not student:
                errors.append({"index": idx, "roll_no": roll, "division": division, "error": "student not found"})
                continue

            res = Result.query.filter_by(batch_id=g.active_batch, roll_no=roll, division=division).first()
            if not res:
                res = Result()
                res.batch_id = g.active_batch
                res.roll_no = roll
                res.division = division
                res.name = student.name
                db.session.add(res)

            # set appropriate grade column
            if subject_code.upper() == "PE":
                res.pe_grade = grade
            elif subject_code.upper() == "EVS":
                res.evs_grade = grade
            else:
                errors.append({"index": idx, "error": "unsupported grade subject"})
                continue

            saved.append({"roll_no": roll, "division": division, "grade": grade})

        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return {"error": "Database error", "details": str(ex)}, 500

    resp = {"message": "Grades saved", "saved": saved}
    if errors:
        resp["errors"] = errors
    return resp, 200


# =====================================================
# Teacher password update
# =====================================================


@teacher_bp.route("/password", methods=["PUT"])
@token_required
def change_password(user_id=None, user_type=None):
    """
    Change password for the logged-in teacher. JSON: { current_password, new_password }
    Requires current password validation.
    """
    if (user_type or "").upper() == "ADMIN":
        return {"error": "Admins should update password via admin flows"}, 403

    data = request.get_json() or {}
    current = data.get("current_password")
    new = data.get("new_password")
    if not current or not new:
        return {"error": "current_password and new_password are required"}, 400

    teacher = Teacher.query.get(user_id)
    if not teacher:
        return {"error": "User not found"}, 404

    if not verify_password(current, teacher.password_hash):
        return {"error": "Current password is incorrect"}, 401

    teacher.password_hash = hash_password(new)
    db.session.commit()

    return {"message": "Password updated"}, 200
