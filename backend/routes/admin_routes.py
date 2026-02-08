# backend/routes/admin_routes.py

from flask import Blueprint, request, jsonify, g

from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, Integer

from app import db
from typing import Any, Dict, Optional, cast
from auth import generate_token
from models import (
    Teacher,
    Subject,
    Student,
    TeacherSubjectAllocation
)
from schemas import StudentSchema
from auth import token_required
from decorators import admin_required
from services.result_service import generate_results_for_division
from batch_config import get_active_batch, set_active_batch
from models import Result, Subject, Mark
from flask import send_file
from io import BytesIO
import os
import json
from datetime import datetime
import math

letter: Optional[Any] = None
canvas_module: Optional[Any] = None
try:
    import importlib
    _rl_pages = importlib.import_module('reportlab.lib.pagesizes')
    _rl_canvas = importlib.import_module('reportlab.pdfgen.canvas')
    letter = getattr(_rl_pages, 'letter', None)
    canvas_module = _rl_canvas
except Exception:
    letter = None
    canvas_module = None

from werkzeug.security import generate_password_hash
from email_utils import send_teacher_credentials_email


admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

student_schema = StudentSchema()

SUBJECT_ORDER = {
    "ENG": 1,
    "HINDI": 2, "IT": 2,
    "ECO": 3,
    "BK": 4,
    "OC": 5,
    "MATHS": 6, "SP": 6,
    "EVS": 7,
    "PE": 8
}


# ======================================================
# 1️⃣ Add Student
# ======================================================
@admin_bp.route("/students", methods=["POST"])
@token_required
@admin_required
def add_student(user_id=None, user_type=None):
    """
    Add a new student (admin only)
    """
    try:
        data = cast(Dict[str, Any], student_schema.load(request.json or {}))
    except Exception as e:
        return {"error": f"Invalid request data: {str(e)}"}, 400

    student = Student(**data)
    # assign to active batch
    try:
        student.batch_id = g.active_batch
    except Exception:
        try:
            student.batch_id = get_active_batch()
        except Exception:
            student.batch_id = None

    try:
        db.session.add(student)
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig).lower() if hasattr(e, 'orig') else str(e)
        if 'unique constraint' in error_msg or 'duplicate' in error_msg:
            return {"error": "Student with this roll number and division already exists"}, 409
        return {"error": f"Failed to add student: {str(e.orig)}"}, 409
    except Exception as ex:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return {"error": f"Database error: {str(ex)}"}, 500

    # Automatically create empty Mark rows for the 4 main/core subjects
    try:
        main_codes = ("ENG", "ECO", "BK", "OC")
        from models import Subject, Mark

        # Pylance may not infer SQLAlchemy column types; ignore attr-type here
        subjects = Subject.query.filter(Subject.subject_code.in_(main_codes)).all()  # type: ignore[attr-defined]
        if subjects is None:
            subjects = []
        for subj in subjects:
            if subj is None or not hasattr(subj, 'subject_id'):
                continue
            # avoid duplicate mark rows if any exist
            exists = Mark.query.filter_by(roll_no=student.roll_no, division=student.division, subject_id=subj.subject_id).first()
            if not exists:
                m = Mark()
                m.roll_no = student.roll_no
                m.division = student.division
                m.subject_id = subj.subject_id
                m.batch_id = student.batch_id
                m.unit1 = 0
                m.unit2 = 0
                m.term = 0
                m.annual = 0
                m.tot = 0
                m.sub_avg = 0
                m.internal = 0
                m.entered_by = None
                db.session.add(m)
        db.session.commit()
    except Exception as mark_err:
        # if this fails, don't block student creation; log the error
        db.session.rollback()
        print(f"Warning: Failed to create default marks for student {student.roll_no}: {str(mark_err)}")

    return {"message": "Student added successfully"}, 201


# ======================================================
# 2️⃣ Get Students (by division)
# ======================================================
@admin_bp.route("/students", methods=["GET"])
@token_required
@admin_required
def list_students(user_id=None, user_type=None):
    """
    List students by division or all students (admin only)
    """
    division = request.args.get("division")
    
    query = Student.query.filter_by(batch_id=g.active_batch)
    if division:
        query = query.filter_by(division=division.upper())
    
    students = (
        query
        .order_by(Student.roll_no)
        .all()
    )

    return jsonify([
        {
            "roll_no": s.roll_no,
            "name": s.name,
            "division": s.division,
            "optional_subject": s.optional_subject,
            "optional_subject_2": s.optional_subject_2
        }
        for s in students
    ]), 200


# ======================================================
# 3️⃣ Assign Teacher to Subject + Division
# ======================================================
@admin_bp.route("/allocations", methods=["POST"])
@token_required
@admin_required
def allocate_teacher(user_id=None, user_type=None):
    """
    Assign teacher to subject & division (admin only)
    """
    data: Dict[str, Any] = (request.json or {})
    teacher_id = data.get("teacher_id")
    subject_id = data.get("subject_id")
    division = data.get("division")

    if not teacher_id or not subject_id or not division:
        return {"error": "teacher_id, subject_id, division required"}, 400

    allocation = TeacherSubjectAllocation()
    allocation.teacher_id = teacher_id
    allocation.subject_id = subject_id
    allocation.division = division

    try:
        db.session.add(allocation)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Allocation already exists"}, 409

    return {"message": "Teacher allocated successfully"}, 201


@admin_bp.route('/allocations/<int:allocation_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_allocation(allocation_id, user_id=None, user_type=None):
    """Delete a specific teacher-subject allocation (admin only)."""
    alloc = TeacherSubjectAllocation.query.get(allocation_id)
    if not alloc:
        return {"error": "Allocation not found"}, 404

    try:
        db.session.delete(alloc)
        db.session.commit()
        return {"message": "Allocation deleted"}, 200
    except Exception as ex:
        db.session.rollback()
        return {"error": "Failed to delete allocation", "details": str(ex)}, 500


# ======================================================
# 4️⃣ View Allocations
# ======================================================
@admin_bp.route("/allocations", methods=["GET"])
@token_required
@admin_required
def list_allocations(user_id=None, user_type=None):
    """
    List all teacher-subject allocations (admin only)
    """
    allocations = TeacherSubjectAllocation.query.all()

    result = []
    for a in allocations:
        teacher = Teacher.query.get(a.teacher_id)
        subj = Subject.query.get(a.subject_id)
        result.append({
            "allocation_id": a.allocation_id,
            "teacher_id": a.teacher_id,
            "teacher_name": teacher.name if teacher else None,
            "subject_id": a.subject_id,
            "subject_code": subj.subject_code if subj else None,
            "subject_name": subj.subject_name if subj else None,
            "division": a.division
        })

    return jsonify(result), 200


# ======================================================
# 5️⃣ Generate Results (per division)
# ======================================================
@admin_bp.route("/results/generate", methods=["POST"])
@token_required
@admin_required
def generate_results(user_id=None, user_type=None):
    """
    Generate / update results for a division (admin only)
    """
    division = (request.json or {}).get("division")
    if not division:
        return {"error": "division is required"}, 400

    generate_results_for_division(division, g.active_batch)

    return {"message": f"Results generated for division {division}"}, 200


# ======================================================
# 6️⃣ Get available divisions (admin)
# ======================================================
@admin_bp.route("/divisions", methods=["GET"])
@token_required
@admin_required
def list_divisions(user_id=None, user_type=None):
    # Ensure active batch is a concrete value for SQLAlchemy comparisons
    try:
        active_batch = getattr(g, 'active_batch', None) or get_active_batch()
    except Exception:
        active_batch = None

    # Query divisions using filter_by
    query = Student.query.with_entities(Student.division).distinct()  # type: ignore[attr-defined]
    if active_batch is not None:
        query = query.filter(Student.batch_id == active_batch)
    divisions = query.all()
    divs = [d[0] for d in divisions]
    return jsonify(divs), 200


@admin_bp.route('/batches', methods=['GET'])
@token_required
@admin_required
def list_batches(user_id=None, user_type=None):
    """Return available batches and currently active batch.

    Prefer the persisted registry.json if present so admins can view
    previously created batches and their metadata.
    """
    reg_path = os.path.join(os.path.dirname(__file__), '..', 'registry.json')
    active = None
    try:
        active = g.active_batch
    except Exception:
        try:
            active = get_active_batch()
        except Exception:
            active = None

    if os.path.exists(reg_path):
        try:
            with open(reg_path, 'r', encoding='utf-8') as f:
                reg = json.load(f)
            batches = reg.get('batches', [])
            return jsonify({"batches": batches, "active_batch": active}), 200
        except Exception:
            pass

    # Fallback: derive batches from existing Student.batch_id values
    batches = Student.query.with_entities(Student.batch_id).distinct().all()  # type: ignore[attr-defined]
    batch_list = [b[0] for b in batches if b[0] is not None]
    return jsonify({"batches": batch_list, "active_batch": active}), 200





@admin_bp.route('/batches/create', methods=['POST'])
@token_required
@admin_required
def create_batch(user_id=None, user_type=None):
    """Create/register a new batch (admin only). This records the batch
    in backend/registry.json and does not modify data; seeding can be
    performed separately via `seed_data.py`.
    """
    data = request.json or {}
    batch_id = data.get('batch_id') or data.get('batch')
    if not batch_id or not str(batch_id).strip():
        return {"error": "batch_id is required"}, 400

    reg_path = os.path.join(os.path.dirname(__file__), '..', 'registry.json')
    try:
        if os.path.exists(reg_path):
            with open(reg_path, 'r', encoding='utf-8') as f:
                reg = json.load(f)
        else:
            reg = {"batches": []}

        batches = reg.get('batches', [])
        # Avoid duplicates
        if any(b.get('batch_id') == batch_id for b in batches):
            return {"error": "batch already exists"}, 409

        entry = {
            "batch_id": str(batch_id),
            "created_at": datetime.utcnow().isoformat(),
            "is_active": False,
            "created_by": user_id or 'admin'
        }
        batches.append(entry)
        reg['batches'] = batches
        with open(reg_path, 'w', encoding='utf-8') as f:
            json.dump(reg, f, indent=2)

        return {"success": True, "batch": entry}, 201
    except Exception as e:
        return {"error": "Failed to register batch", "details": str(e)}, 500


@admin_bp.route('/batches/switch', methods=['POST'])
@token_required
@admin_required
def switch_batch_with_registry(user_id=None, user_type=None):
    """Switch active batch and update registry.json to mark active batch."""
    data = request.json or {}
    batch_id = data.get('batch_id')
    if not batch_id:
        return {"error": "batch_id is required"}, 400

    try:
        # Persist active batch
        set_active_batch(batch_id)

        # Update registry.json active flags if present
        reg_path = os.path.join(os.path.dirname(__file__), '..', 'registry.json')
        if os.path.exists(reg_path):
            with open(reg_path, 'r', encoding='utf-8') as f:
                reg = json.load(f)
            changed = False
            for b in reg.get('batches', []):
                if b.get('batch_id') == batch_id:
                    if not b.get('is_active'):
                        b['is_active'] = True
                        changed = True
                else:
                    if b.get('is_active'):
                        b['is_active'] = False
                        changed = True
            if changed:
                with open(reg_path, 'w', encoding='utf-8') as f:
                    json.dump(reg, f, indent=2)

        return {"message": f"Active batch set to {batch_id}"}, 200
    except Exception as ex:
        return {"error": "Failed to set active batch", "details": str(ex)}, 500


# ======================================================
# 7️⃣ Fetch results by division or roll_no (admin)
# Query params: division OR roll_no (+ optional division)
# ======================================================
@admin_bp.route("/results", methods=["GET"])
@token_required
@admin_required
def fetch_results(user_id=None, user_type=None):
    roll_no = request.args.get("roll_no")
    division = request.args.get("division")
    # function-scoped default to avoid UnboundLocalError when different branches
    # conditionally set this variable later.
    excel_marks = None

    # If roll_no provided, optionally restrict by division
    if roll_no:
        students = Student.query.filter_by(roll_no=roll_no, batch_id=g.active_batch)
        if division:
            students = students.filter_by(division=division)
        students = students.all()
        if not students:
            return {"error": "Student not found"}, 404

        # Ensure results are generated for involved divisions
        for s in students:
            try:
                generate_results_for_division(s.division, g.active_batch)
            except Exception:
                pass

        # Build rows for each matching student (usually one)
        rows = []
        # Prefer master Excel file data if available
        from config import MASTER_EXCEL_PATH, MASTER_EXCEL_SHEET
        use_excel = False
        try:
            import os
            if os.path.exists(MASTER_EXCEL_PATH):
                use_excel = True
        except Exception:
            use_excel = False
        for s in students:
            # If a master Excel exists, try to read detailed marks from it for this roll
            excel_marks = None
            if use_excel:
                try:
                    import openpyxl, io
                    data = openpyxl.load_workbook(MASTER_EXCEL_PATH, data_only=True)
                    if MASTER_EXCEL_SHEET in data.sheetnames:
                        sh = data[MASTER_EXCEL_SHEET]
                        # build header map
                        it = sh.iter_rows(values_only=True)
                        headers = [str(x).strip().lower() if x is not None else '' for x in next(it)]
                        def idx_of(names):
                            for n in names:
                                if n in headers:
                                    return headers.index(n)
                            return None
                        r_idx = idx_of(['roll_no', 'roll', 'rollno'])
                        d_idx = idx_of(['division', 'div'])
                        # subject column may be subject code
                        subj_idx = idx_of(['subject', 'subject_code', 'subject_id'])
                        u1_idx = idx_of(['unit1'])
                        u2_idx = idx_of(['unit2'])
                        term_idx = idx_of(['term'])
                        annual_idx = idx_of(['annual'])
                        grace_idx = idx_of(['internal', 'grace'])
                        for row in it:
                            if not row or all(c is None for c in row):
                                continue
                            rv = row[r_idx] if r_idx is not None and r_idx < len(row) else None
                            dv = row[d_idx] if d_idx is not None and d_idx < len(row) else None
                            if rv is None:
                                continue
                            if str(rv).strip() == str(s.roll_no).strip() and (not division or (dv and str(dv).strip() == str(s.division).strip())):
                                # found matching excel row
                                excel_marks = {
                                    'unit1': row[u1_idx] if u1_idx is not None and u1_idx < len(row) else None,
                                    'unit2': row[u2_idx] if u2_idx is not None and u2_idx < len(row) else None,
                                    'term': row[term_idx] if term_idx is not None and term_idx < len(row) else None,
                                    'annual': row[annual_idx] if annual_idx is not None and annual_idx < len(row) else None,
                                    'internal': row[grace_idx] if grace_idx is not None and grace_idx < len(row) else None,
                                }
                                break
                except Exception:
                    excel_marks = None

            result = Result.query.filter_by(roll_no=s.roll_no, division=s.division, batch_id=g.active_batch).first()
            # If result is missing, fall back to available Marks so UI can show partial data
            marks = Mark.query.filter_by(roll_no=s.roll_no, division=s.division, batch_id=g.active_batch).all()
            if marks is None:
                marks = []
            
            # Build subjects map with validation
            all_subjects_list = Subject.query.all()
            if all_subjects_list is None:
                all_subjects_list = []
            subjects_map = {}
            for sub in all_subjects_list:
                if sub is not None and hasattr(sub, 'subject_id') and hasattr(sub, 'subject_code'):
                    subjects_map[sub.subject_id] = sub.subject_code
            
            mark_map = {}
            for m in marks:
                if m is None or not hasattr(m, 'subject_id'):
                    continue
                code = subjects_map.get(m.subject_id)
                if code is not None:  # Explicit None check before using code
                    mark_map[code] = m

            # If a master Excel was used in the roll_no path, it's local there; for division path
            # ensure excel_marks is defined to avoid UnboundLocalError. We do not auto-load
            # the master Excel for whole-division queries here (keeps behavior simple).
            excel_marks = None
            subject_entries = []
            total_avg = 0
            total_grace = 0

            # Build the list of subject codes to display: core subjects, any optional subjects, and any codes present in marks
            all_subjects = Subject.query.filter_by(active=True).all()
            include_codes = set()
            for sub in all_subjects:
                if sub.subject_type == 'CORE':
                    include_codes.add(sub.subject_code)
            if s.optional_subject:
                include_codes.add(s.optional_subject)
            if s.optional_subject_2:
                include_codes.add(s.optional_subject_2)
            # include any codes present in mark rows to avoid missing any entered subjects
            include_codes.update(mark_map.keys())

            # Custom sort function
            def get_subject_order(code):
                return SUBJECT_ORDER.get(code, 99)

            # iterate in sorted order
            for code in sorted(include_codes, key=get_subject_order):
                # handle grade-only subjects separately
                if code in ("EVS", "PE"):
                    # Check Result first
                    res_grade = getattr(result, f"{code.lower()}_grade", None) if result else None
                    if res_grade:
                        # Found in Result
                        subject_entries.append({
                             "code": code,
                             "grade": res_grade,
                             "mark": {
                                 "unit1": None, "unit2": None, "term": None, "annual": None,
                                 "internal": 0, "sub_avg": 0, "tot": 0, "grace": 0
                             }
                        })
                    else:
                        # Fallback to Marks table if any
                        m = mark_map.get(code)
                        if m and m.annual is not None:
                            a = m.annual
                            grade = None
                            if a >= 75: grade = 'A+'
                            elif a >= 60: grade = 'A'
                            elif a >= 50: grade = 'B'
                            elif a >= 35: grade = 'C'
                            else: grade = 'F'
                            
                            subject_entries.append({
                                "code": code, 
                                "grade": grade, 
                                "mark": {
                                    "annual": m.annual, 
                                    "mark_id": m.mark_id, 
                                    "unit1": m.unit1, 
                                    "unit2": m.unit2, 
                                    "term": m.term, 
                                    "tot": m.tot, 
                                    "sub_avg": m.sub_avg, 
                                    "internal": getattr(m, 'internal', 0),
                                    "grace": getattr(result, f"{code.lower()}_grace", 0) if result else 0
                                }
                            })
                        else:
                            # Force entry even if no data (as empty grade)
                            subject_entries.append({
                                "code": code,
                                "grade": "-",
                                "mark": {
                                    "unit1": None, "unit2": None, "term": None, "annual": None,
                                    "internal": 0, "sub_avg": 0, "tot": 0, "grace": 0
                                }
                            })
                    continue

                avg, grace = (None, 0.0)
                if result:
                    avg, grace = result.get_subject_data(code)
                else:
                    m = mark_map.get(code)
                    avg = m.annual if m and m.annual is not None else None
                    grace = getattr(m, 'internal', 0) if m else 0

                final = None
                if avg is not None:
                    final = (avg or 0) + (grace or 0)
                    total_avg += avg or 0
                    total_grace += grace or 0

                # include detailed mark breakdown if available; prefer excel row values when present
                m = mark_map.get(code)
                mark_detail = None
                if excel_marks is not None and isinstance(excel_marks, dict):
                    # order: unit1, term, unit2, internal, annual, grace, total
                    try:
                        total_val = None
                        u1 = float(excel_marks.get('unit1') or 0)
                        u2 = float(excel_marks.get('unit2') or 0)
                        term_v = float(excel_marks.get('term') or 0)
                        annual_v = float(excel_marks.get('annual') or 0)
                        total_val = u1 + u2 + term_v + annual_v
                    except (TypeError, ValueError):
                        total_val = None
                    mark_detail = {
                        "unit1": excel_marks.get('unit1'),
                        "term": excel_marks.get('term'),
                        "unit2": excel_marks.get('unit2'),
                        "internal": excel_marks.get('internal'),
                        "annual": excel_marks.get('annual'),
                        "grace": None,
                        "total": total_val,
                    }
                elif m is not None:
                    total_val = m.tot if m.tot is not None else ((m.unit1 or 0) + (m.unit2 or 0) + (m.term or 0) + (m.annual or 0))
                    mark_detail = {
                        "unit1": m.unit1,
                        "term": m.term,
                        "unit2": m.unit2,
                        "internal": getattr(m, 'internal', 0),
                        "annual": m.annual,
                        "grace": grace,
                        "total": total_val,
                    }

                subject_entries.append({
                    "code": code,
                    "avg": avg,
                    "grace": grace,
                    "final": final,
                    "mark": mark_detail if mark_detail is not None else {"unit1": None, "unit2": None, "term": None, "annual": None, "internal": 0, "sub_avg": 0, "tot": 0, "grace": 0}
                })
                
                # Ensure mark_detail has sub_avg if not present
                if subject_entries[-1]["mark"]:
                    if "sub_avg" not in subject_entries[-1]["mark"]:
                        # try generic
                        if m is not None:
                            subject_entries[-1]["mark"]["sub_avg"] = m.sub_avg if hasattr(m, 'sub_avg') else 0
                        else:
                            subject_entries[-1]["mark"]["sub_avg"] = 0

            final_total = None
            if subject_entries:
                # Only show final_total if percentage exists (i.e., result fully computed)
                perc = getattr(result, "percentage", None) if result else None
                if perc is not None:
                    final_total = total_avg + total_grace

            rows.append({
                "roll_no": s.roll_no,
                "name": s.name,
                "division": s.division,
                "subjects": subject_entries,
                "total_avg": round(total_avg, 2),
                "total_grace": round(total_grace, 2),
                "final_total": round(final_total, 2) if final_total is not None else None,
                "percentage": getattr(result, "percentage", None) if result else None,
                "overall_grade": getattr(result, "overall_grade", None) if result else None,
            })

        # If caller requested a single roll_no, return single object
        if len(rows) == 1:
            return jsonify(rows[0]), 200
        return jsonify(rows), 200

    # Else, require division
    if not division:
        return {"error": "division or roll_no is required"}, 400

    # regenerate results for the division
    try:
        generate_results_for_division(division, g.active_batch)
    except Exception:
        pass

    # Build rows for entire division
    students = Student.query.filter_by(division=division, batch_id=g.active_batch).order_by(Student.roll_no).all()
    if students is None:
        students = []
    rows = []
    for idx, s in enumerate(students, start=1):
        if s is None:
            continue
        result = Result.query.filter_by(roll_no=s.roll_no, division=s.division, batch_id=g.active_batch).first()
        # Prepare mark map to allow partial display when Result row missing
        marks = Mark.query.filter_by(roll_no=s.roll_no, division=s.division, batch_id=g.active_batch).all()
        if marks is None:
            marks = []
        
        # Build subjects map with validation
        all_subjects_list = Subject.query.all()
        if all_subjects_list is None:
            all_subjects_list = []
        subjects_map = {}
        for sub in all_subjects_list:
            if sub is not None and hasattr(sub, 'subject_id') and hasattr(sub, 'subject_code'):
                subjects_map[sub.subject_id] = sub.subject_code
        
        mark_map = {}
        for m in marks:
            if m is None or not hasattr(m, 'subject_id'):
                continue
            code = subjects_map.get(m.subject_id)
            if code is not None:  # Explicit None check before using code
                mark_map[code] = m

        subject_entries = []
        total_avg = 0
        total_grace = 0

        # Build display set: core subjects + student's optionals + any subjects present in marks
        all_subjects = Subject.query.filter_by(active=True).all()
        include_codes = set()
        for sub in all_subjects:
            if sub.subject_type == 'CORE':
                include_codes.add(sub.subject_code)
        if s.optional_subject:
            include_codes.add(s.optional_subject)
        if s.optional_subject_2:
            include_codes.add(s.optional_subject_2)
        include_codes.update(mark_map.keys())

        # Custom sort function
        def get_subject_order(code):
            return SUBJECT_ORDER.get(code, 99)

        for code in sorted(include_codes, key=get_subject_order):
            # grade-only subjects
            if code in ("EVS", "PE"):
                # Check Result first
                res_grade = getattr(result, f"{code.lower()}_grade", None) if result else None
                if res_grade:
                     subject_entries.append({
                         "code": code, 
                         "grade": res_grade,
                         "mark": {
                             "unit1": None, "unit2": None, "term": None, "annual": None,
                             "internal": 0, "sub_avg": 0, "tot": 0, "grace": 0
                         }
                     })
                else:
                     m = mark_map.get(code)
                     if m and m.annual is not None:
                        a = m.annual
                        grade = None
                        if a >= 75: grade = 'A+'
                        elif a >= 60: grade = 'A'
                        elif a >= 50: grade = 'B'
                        elif a >= 35: grade = 'C'
                        else: grade = 'F'
                        subject_entries.append({
                            "code": code, 
                            "grade": grade, 
                            "mark": {
                                "unit1": m.unit1, 
                                "term": m.term, 
                                "unit2": m.unit2, 
                                "internal": getattr(m, 'internal', 0), 
                                "annual": m.annual, 
                                "grace": getattr(result, f"{code.lower()}_grace", 0) if result else getattr(m, 'internal', 0), 
                                "total": m.tot,
                                "sub_avg": m.sub_avg
                            }
                        })
                     else:
                        # Force entry even if no data (as empty grade)
                        subject_entries.append({
                            "code": code,
                            "grade": "-",
                            "mark": {
                                "unit1": None, "unit2": None, "term": None, "annual": None,
                                "internal": 0, "sub_avg": 0, "tot": 0, "grace": 0
                            }
                        })
                continue

            avg, grace = (None, 0.0)
            if result:
                avg, grace = result.get_subject_data(code)
            else:
                m = mark_map.get(code)
                avg = m.annual if m and m.annual is not None else None
                grace = getattr(m, 'internal', 0) if m else 0

            final = None
            if avg is not None:
                final = (avg or 0) + (grace or 0)
                total_avg += avg or 0
                total_grace += grace or 0

            m = mark_map.get(code)
            mark_detail = None
            if m:
                total_val = m.tot if m.tot is not None else ((m.unit1 or 0) + (m.unit2 or 0) + (m.term or 0) + (m.annual or 0))
                mark_detail = {
                    "unit1": m.unit1,
                    "term": m.term,
                    "unit2": m.unit2,
                    "internal": getattr(m, 'internal', 0),
                    "annual": m.annual,
                    "grace": grace,
                    "total": total_val,
                }
            elif excel_marks is not None:
                # excel marks available but no DB mark row
                m_internal = excel_marks.get('internal')
                total_val = None
                try:
                    u1 = float(excel_marks.get('unit1') or 0)
                    u2 = float(excel_marks.get('unit2') or 0)
                    term_v = float(excel_marks.get('term') or 0)
                    annual_v = float(excel_marks.get('annual') or 0)
                    total_val = u1 + u2 + term_v + annual_v
                except Exception:
                    total_val = None
                mark_detail = {
                    "unit1": excel_marks.get('unit1'),
                    "term": excel_marks.get('term'),
                    "unit2": excel_marks.get('unit2'),
                    "internal": m_internal,
                    "annual": excel_marks.get('annual'),
                    "grace": None,
                    "total": total_val,
                }

            subject_entries.append({"code": code, "avg": avg, "grace": grace, "final": final, "mark": mark_detail})

            # Ensure mark_detail has sub_avg if not present
            if subject_entries[-1]["mark"] and "sub_avg" not in subject_entries[-1]["mark"]:
                     # try generic
                     if m: subject_entries[-1]["mark"]["sub_avg"] = m.sub_avg
                     else: subject_entries[-1]["mark"]["sub_avg"] = 0

        final_total = None
        if subject_entries:
            perc = getattr(result, "percentage", None) if result else None
            if perc is not None:
                final_total = total_avg + total_grace

        rows.append({
            "seq": idx,
            "roll_no": s.roll_no,
            "name": s.name,
            "subjects": subject_entries,
            "total_avg": round(total_avg, 2),
            "total_grace": round(total_grace, 2),
            "final_total": round(final_total, 2) if final_total is not None else None,
            "percentage": getattr(result, "percentage", None) if result else None,
            "overall_grade": getattr(result, "overall_grade", None) if result else None,
        })

    return jsonify(rows), 200







# ======================================================
# Download student marksheet PDF (admin only)
# ======================================================
@admin_bp.route('/students/<string:roll_no>/pdf', methods=['GET'])
@token_required
@admin_required
def student_marksheet_pdf(roll_no, user_id=None, user_type=None):
    division = request.args.get('division')
    if not division:
        return {"error": "division is required"}, 400

    # ensure results are up-to-date
    try:
        generate_results_for_division(division, g.active_batch)
    except Exception:
        pass

    res = Result.query.filter_by(roll_no=roll_no, division=division, batch_id=g.active_batch).first()
    if not res:
        return {"error": "Result not found"}, 404

    if canvas_module is None or letter is None:
        return {"error": "reportlab not installed on server. Install reportlab in requirements."}, 501

    buf = BytesIO()
    # canvas_module is the imported reportlab.pdfgen.canvas module
    CanvasClass = getattr(canvas_module, 'Canvas', None)
    if CanvasClass is None:
        return {"error": "reportlab canvas API not available"}, 501
    c = CanvasClass(buf, pagesize=cast(Any, letter))
    width, height = cast(Any, letter)

    # Header
    c.setFont('Helvetica-Bold', 16)
    c.drawString(40, height - 50, 'Official Marksheet')
    c.setFont('Helvetica', 12)
    c.drawString(40, height - 70, f'Name: {res.name}  |  Roll: {res.roll_no}  |  Division: {res.division}')

    # Table header
    y = height - 110
    c.setFont('Helvetica-Bold', 11)
    c.drawString(40, y, 'Subject')
    c.drawString(260, y, 'Annual')
    c.drawString(360, y, 'Internal')
    c.drawString(460, y, 'Final')
    c.setFont('Helvetica', 11)
    y -= 18

    # iterate known subjects and map to result fields
    mapping = [
        ('ENG', 'eng_avg', 'eng_grace'),
        ('ECO', 'eco_avg', 'eco_grace'),
        ('BK', 'bk_avg', 'bk_grace'),
        ('OC', 'oc_avg', 'oc_grace'),
        ('HINDI', 'hindi_avg', 'hindi_grace'),
        ('IT', 'it_avg', 'it_grace'),
        ('MATHS', 'maths_avg', 'maths_grace'),
        ('SP', 'sp_avg', 'sp_grace')
    ]

    total = 0
    for code, avg_field, grace_field in mapping:
        avg = getattr(res, avg_field, None)
        grace = getattr(res, grace_field, 0) or 0
        if avg is None:
            continue
        final = (avg or 0) + (grace or 0)
        c.drawString(40, y, code)
        c.drawRightString(320, y, f'{round(avg,2)}')
        c.drawRightString(420, y, f'{round(grace,2)}')
        c.drawRightString(520, y, f'{round(final,2)}')
        total += final
        y -= 16

    y -= 8
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, f'Total: {round(res.percentage * (len(mapping) / len(mapping)),2) if res.percentage is not None else "-"}')
    c.drawRightString(520, y, f'Percentage: {res.percentage or "-"}')

    c.showPage()
    c.save()
    buf.seek(0)

    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=f'{roll_no}_marksheet.pdf')




# ======================================================
# ADMIN – TEACHERS CRUD
# ======================================================

@admin_bp.route("/teachers", methods=["GET"])
@token_required
def list_teachers(user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    teachers = Teacher.query.all()
    return jsonify([
        {
            "teacher_id": t.teacher_id,
            "name": t.name,
            "userid": t.userid,
            "email": t.email,
            "active": t.active,
            "role": t.role
        }
        for t in teachers
    ]), 200


@admin_bp.route("/teachers", methods=["POST"])
@token_required
def add_teacher(user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    data = request.json or {}
    required = ["name", "userid", "password"]
    if not all(k in data for k in required):
        return {"error": "Missing required fields: name, userid, password"}, 400

    # Check for duplicate userid
    if Teacher.query.filter_by(userid=data["userid"]).first():
        return {"error": "UserID already exists"}, 409

    try:
        # Instantiate without calling the model constructor with unknown kwargs
        teacher = Teacher()
        teacher.name = data["name"]
        teacher.userid = data["userid"]
        teacher.email = data.get("email")
        teacher.role = data.get("role", "TEACHER")
        teacher.active = True
        teacher.password_hash = generate_password_hash(data["password"])

        db.session.add(teacher)
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig).lower() if hasattr(e, 'orig') else str(e)
        if 'unique constraint' in error_msg:
            return {"error": "This UserID already exists"}, 409
        return {"error": f"Database constraint violation: {str(e.orig)}"}, 409
    except Exception as ex:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to add teacher: {str(ex)}"}, 500

    # Send email notification if email is provided (non-blocking)
    email_sent = False
    email_error = None
    if teacher.email:
        try:
            send_teacher_credentials_email(
                teacher_name=teacher.name,
                teacher_email=teacher.email,
                username=teacher.userid,
                password=data["password"]  # Send plain text password
            )
            email_sent = True
        except Exception as e:
            # Log the error but don't fail the teacher creation
            email_error = str(e)
            print(f"[WARNING] Failed to send email to {teacher.email}: {email_error}")

    response = {"message": "Teacher added successfully", "teacher_id": teacher.teacher_id}
    if email_error:
        response["email_warning"] = f"Teacher created but email notification failed: {email_error[:100]}"
    
    return response, 201



@admin_bp.route("/teachers/<int:teacher_id>", methods=["PUT"])
@token_required
def update_teacher(teacher_id, user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    teacher = Teacher.query.get(teacher_id)
    if not teacher:
        return {"error": "Teacher not found"}, 404

    data = request.json or {}

    try:
        teacher.name = data.get("name", teacher.name)
        teacher.userid = data.get("userid", teacher.userid)
        teacher.email = data.get("email", teacher.email)
        teacher.active = data.get("active", teacher.active)

        # Only update password when a non-empty new password is provided.
        # This prevents overwriting the stored hashed password when the
        # frontend leaves the password field blank (meaning "keep old").
        password = (data.get("password") or "").strip()
        password_updated = False
        if password:
            teacher.password_hash = generate_password_hash(password)
            password_updated = True

        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig).lower() if hasattr(e, 'orig') else str(e)
        if 'unique constraint' in error_msg:
            return {"error": "UserID already exists"}, 409
        return {"error": f"Database constraint violation: {str(e.orig)}"}, 409
    except Exception as ex:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to update teacher: {str(ex)}"}, 500

    # Send email notification if password was updated and email is provided (non-blocking)
    email_error = None
    if password_updated and teacher.email:
        try:
            send_teacher_credentials_email(
                teacher_name=teacher.name,
                teacher_email=teacher.email,
                username=teacher.userid,
                password=password  # Send the new plain text password
            )
        except Exception as e:
            # Log the error but don't fail the teacher update
            email_error = str(e)
            print(f"[WARNING] Failed to send email to {teacher.email}: {email_error}")

    response = {"message": "Teacher updated"}
    if email_error:
        response["email_warning"] = f"Teacher updated but email notification failed: {email_error[:100]}"
    
    return response, 200



@admin_bp.route("/teachers/<int:teacher_id>", methods=["DELETE"])
@token_required
def delete_teacher(teacher_id, user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    teacher = Teacher.query.get(teacher_id)
    if not teacher:
        return {"error": "Teacher not found"}, 404

    try:
        db.session.delete(teacher)
        db.session.commit()
        return {"message": "Teacher deleted successfully"}, 200
    except IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig).lower() if hasattr(e, 'orig') else str(e)
        if 'foreign key constraint' in error_msg or 'fk_' in error_msg:
            return {"error": "Cannot delete teacher: they have active allocations or grades. Please unassign subjects first."}, 409
        return {"error": f"Database constraint violation: {str(e.orig)}"}, 409
    except Exception as ex:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to delete teacher: {str(ex)}"}, 500


# ======================================================
# ADMIN LOGIN (tests expect /admin/login)
# ======================================================
@admin_bp.route("/login", methods=["POST"])
def admin_login():
    data = request.json or {}
    userid = data.get("userid")
    password = data.get("password")

    if not userid or not password:
        return {"error": "userid and password required"}, 400

    admin = None
    from models import Admin
    admin = Admin.query.filter_by(username=userid, active=True).first()
    if not admin:
        return {"error": "Invalid credentials"}, 401

    # Admin passwords are hashed with werkzeug.generate_password_hash
    from werkzeug.security import check_password_hash
    if not check_password_hash(admin.password_hash, password):
        return {"error": "Invalid credentials"}, 401

    token = generate_token(admin.admin_id, "ADMIN")
    # Return role in lowercase for consistency with client/tests
    return {"token": token, "role": "admin"}, 200


# ======================================================
# ADMIN IMPERSONATE TEACHER (open teacher panel without logging out)
# ======================================================
@admin_bp.route('/teachers/<int:teacher_id>/impersonate', methods=['POST'])
@token_required
@admin_required
def impersonate_teacher(teacher_id, user_id=None, user_type=None):
    """
    Generate a short-lived token for a teacher so an admin can open
    the teacher panel without logging out (impersonation).
    """
    teacher = Teacher.query.get(teacher_id)
    if not teacher or not teacher.active:
        return {"error": "Teacher not found or inactive"}, 404

    # Issue token with TEACHER role
    token = generate_token(teacher.teacher_id, "TEACHER", expires_hours=2)

    return {
        "token": token,
        "teacher": {
            "teacher_id": teacher.teacher_id,
            "name": teacher.name,
            "userid": teacher.userid,
            "email": teacher.email,
        }
    }, 200

#======================================================
# 1.5️⃣ Import Students from Excel
# ======================================================
@admin_bp.route("/students/import", methods=["POST"])
@token_required
@admin_required
def import_students(user_id=None, user_type=None):
    """
    Import students from Excel file (admin only)
    Excel can have multiple sheets, each with columns: Roll Number, Name, Division, Optional Subject 1, Optional Subject 2
    """
    print("[IMPORT] Starting import request")
    
    if 'file' not in request.files:
        print("[IMPORT] No file in request.files")
        return {"error": "No file provided"}, 400

    file = request.files['file']
    print(f"[IMPORT] File received: {file.filename}")
    
    if file.filename == '':
        print("[IMPORT] Empty filename")
        return {"error": "No file selected"}, 400

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        print(f"[IMPORT] Invalid file extension: {file.filename}")
        return {"error": "File must be Excel (.xlsx or .xls)"}, 400

    try:
        import openpyxl
        print("[IMPORT] Loading workbook")
        wb = openpyxl.load_workbook(file.stream, data_only=True)
        print(f"[IMPORT] Workbook loaded with sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"[IMPORT] Failed to read Excel file: {str(e)}")
        # Return details as an array for consistent client handling
        return {"error": "Failed to read Excel file", "details": [str(e)]}, 400

    if not wb.sheetnames:
        print("[IMPORT] No sheets in workbook")
        return {"error": "Excel file contains no sheets"}, 400

    students_created = 0
    errors = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        print(f"[IMPORT] Processing sheet '{sheet_name}' with {len(rows)} rows")
        
        if not rows:
            continue
        
        # Skip header row
        data_rows = rows[1:]
        
        for row_idx, row in enumerate(data_rows, start=2):  # start=2 because Excel rows are 1-indexed, and we skipped header
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            try:
                # Map columns: Roll Number (0), Name (1), Division (2), Optional Subject 1 (3), Optional Subject 2 (4)
                roll_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else None
                name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
                division = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
                optional_subject = str(row[3]).strip() if len(row) > 3 and row[3] is not None and str(row[3]).strip() != '' else None
                optional_subject_2 = str(row[4]).strip() if len(row) > 4 and row[4] is not None and str(row[4]).strip() != '' else None
                
                print(f"[IMPORT] Processing row {row_idx}: roll_no={roll_no}, name={name}, division={division}")
                
                if not roll_no or not name or not division:
                    errors.append(f"Sheet '{sheet_name}', Row {row_idx}: Missing required fields (Roll Number, Name, Division)")
                    continue
                
                # Create student and assign to active batch
                student = Student()
                student.roll_no = roll_no
                student.name = name
                student.division = division.upper()
                student.optional_subject = optional_subject
                student.optional_subject_2 = optional_subject_2
                try:
                    student.batch_id = g.active_batch
                except Exception:
                    try:
                        student.batch_id = get_active_batch()
                    except Exception:
                        student.batch_id = None
                
                try:
                    db.session.add(student)
                    db.session.commit()
                    students_created += 1
                    
                    print(f"[IMPORT] Created student: {roll_no} ({name}) in {division}")
                    
                    # Automatically create empty Mark rows for the 4 main/core subjects
                    main_codes = ("ENG", "ECO", "BK", "OC")
                    subjects = Subject.query.filter(Subject.subject_code.in_(main_codes)).all()  # type: ignore[attr-defined]
                    if subjects is None:
                        subjects = []
                    for subj in subjects:
                        if subj is None or not hasattr(subj, 'subject_id'):
                            continue
                        # avoid duplicate mark rows if any exist
                        exists = Mark.query.filter_by(roll_no=student.roll_no, division=student.division, subject_id=subj.subject_id).first()
                        if not exists:
                            m = Mark()
                            m.roll_no = student.roll_no
                            m.division = student.division
                            m.subject_id = subj.subject_id
                            m.batch_id = student.batch_id
                            m.unit1 = 0
                            m.unit2 = 0
                            m.term = 0
                            m.annual = 0
                            m.tot = 0
                            m.sub_avg = 0
                            m.internal = 0
                            m.entered_by = None
                            db.session.add(m)
                    db.session.commit()
                    
                except IntegrityError:
                    db.session.rollback()
                    errors.append(f"Sheet '{sheet_name}', Row {row_idx}: Student {roll_no} in division {division} already exists")
                except Exception as e:
                    db.session.rollback()
                    errors.append(f"Sheet '{sheet_name}', Row {row_idx}: Failed to create student - {str(e)}")
                    
            except Exception as e:
                errors.append(f"Sheet '{sheet_name}', Row {row_idx}: Error processing row - {str(e)}")

    print(f"[IMPORT] Import completed: {students_created} students created, {len(errors)} errors")
    
    if students_created == 0 and errors:
        return {"error": "No students imported", "details": errors}, 400
    
    response = {"message": f"Successfully imported {students_created} students"}
    if errors:
        response["warnings"] = errors  # type: ignore
    
    print(f"[IMPORT] Returning response: {response}")
    return response, 201


# ======================================================
# EXPORT RESULTS TO EXCEL
# ======================================================
@admin_bp.route("/results/export-excel", methods=["GET"])
@token_required
@admin_required
def export_results_to_excel(user_id=None, user_type=None):
    """
    Export all students' complete results for the active BATCH to Excel.
    
    Generates a multi-sheet workbook for the entire batch.
    Includes per-division sheets and a consolidated summary.
    """
    # Division param is no longer strictly required for generation, 
    # but we might still accept it if we wanted to filter (though requirement is Batch Wide).
    # For now, we ignore division param or use it just for logging.
    
    try:
        # Import Excel export utility
        from services.excel_export import generate_excel_for_batch
        
        batch_id = g.active_batch
        if not batch_id:
             return {"error": "No active batch found"}, 400

        # Generate the workbook
        wb = generate_excel_for_batch(batch_id)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        # Format: FYJC Result {batch_id}.xlsx
        # We might want to sanitize batch_id to be safe for filenames
        safe_batch = str(batch_id).replace("/", "-").replace("\\", "-")
        filename = f"FYJC Result {safe_batch}.xlsx"
        
        # Return as file download
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"[EXCEL EXPORT ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to generate Excel: {str(e)}"}, 500