from typing import Optional, Dict, List, Set, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from models import Student, Result, Mark, Subject
from app import db


def generate_excel_for_batch(batch_id: str) -> Workbook:
    """
    Generates a multi-sheet Excel workbook for the entire batch.
    Sheets per Division:
      - DIV_{X}: Detailed Marksheet (Units, Terms, etc.)
      - DIV_{X}_All: Summary Ledger
      - NAME {X}: Name List
    Global Sheets:
      - Sheet2: Consolidated Summary (All Divisions)
    """
    wb = Workbook()
    
    # Remove default sheet
    default_ws = wb.active
    if default_ws:
        wb.remove(default_ws)

    # ---------------- STYLES ----------------
    # Dark Blue Header (Main Sheets)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Light Blue Sub-Header
    sub_header_fill = PatternFill("solid", fgColor="D9E1F2")
    sub_header_font = Font(bold=True, size=9)

    # Summary Sheet Styles
    summary_header_fill = PatternFill("solid", fgColor="BDD7EE") # Lighter blue for Sheet2
    summary_header_font = Font(bold=True, size=10)

    # Common
    data_font = Font(size=9)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ---------------- DATA FETCHING ----------------
    students = Student.query.filter_by(batch_id=batch_id).order_by(Student.division, Student.roll_no).all()
    if not students:
        # Create empty debug sheet if no data
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No students found for this batch."
        return wb

    divisions = sorted(list(set(s.division for s in students)))

    import re

    def natural_sort_key(s):
        """Standard natural sort key for roll numbers like 1, 2, 10."""
        return [int(text) if text.isdigit() else text.lower()
                for text in re.split('([0-9]+)', s.roll_no)]

    # Sort ALL students first, so grouping by division keeps order if we iterate
    # Actually, we group by division then sort inside division
    # But let's keep consistent Sort
    students.sort(key=natural_sort_key)

    marks = Mark.query.filter_by(batch_id=batch_id).all()
    results = Result.query.filter_by(batch_id=batch_id).all()
    all_subjects = Subject.query.all()
    
    # Mappings
    # Mark Map: (roll_no, subject_code) -> Mark Object (for Units/Term breakdown)
    subject_id_code_map = {s.subject_id: s.subject_code for s in all_subjects}
    mark_map = {(m.roll_no, subject_id_code_map.get(m.subject_id)): m for m in marks}
    
    # Result Map: roll_no -> Result Object (for Averages, Grace, Totals)
    result_map: Dict[str, Result] = {res.roll_no: res for res in results}

    # Student Division Grouping
    students_by_div = {d: [] for d in divisions}
    for s in students:
        students_by_div[s.division].append(s)


    # ---------------- HELPER: COLUMN WIDTH ----------------
    def autosize_columns(ws, data_start_row=1):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip header rows for width calculation if overly verbose, usually standard is fine
                if cell.row < data_start_row: continue 
                
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # Cap width
            ws.column_dimensions[col_letter].width = min(adjusted_width, 40)


    # ---------------- GENERATE SHEETS ----------------

    # We will collect summary rows by division for Sheet2
    all_summary_rows_by_div = {} 

    # Static summary headers used in DIV_{X}_All and consolidated sheets
    summary_headers = [
        "ROLL NO", "STUDENT NAME",
        "ENG", "Eng Grace",
        "HINDI / IT", "Opt1 Grace",
        "ECO", "Eco Grace",
        "BK", "Bk Grace",
        "OC", "Oc Grace",
        "MATHS / SP", "Opt2 Grace",
        "TOTAL", "Grace Total", "Per.", "Result"
    ]

    for div in divisions:
        div_students = students_by_div[div]

        # ==========================================
        # 1. DIV_{X} (Detailed Marksheet)
        # ==========================================
        ws_detail = wb.create_sheet(f"DIV_{div}")
        
        # --- Titles ---
        ws_detail.merge_cells("A1:AZ1")
        ws_detail["A1"] = "SIES JUNIOR COLLEGE OF COMMERCE, NERUL"
        ws_detail["A1"].font = Font(bold=True, size=14)
        ws_detail["A1"].alignment = center_align

        ws_detail.merge_cells("A2:AZ2")
        ws_detail["A2"] = f"FYJC ( DIV {div} ) MARKSHEET – {batch_id}"
        ws_detail["A2"].font = Font(bold=True, size=12)
        ws_detail["A2"].alignment = center_align

        ws_detail.merge_cells("A3:AZ3")
        ws_detail["A3"] = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws_detail["A3"].alignment = center_align

        # --- Headers ---
        # Fixed Order: ENG, ECO, BK, OC, OPT1, OPT2, EVS, PE
        # Display Names: "ENGLISH", "ECONOMICS", "BOOK KEEPING", "ORG OF COMM", "HINDI / IT", "MATHS / SP", "EVS", "PE"
        
        detail_cols_struct = [
            ("ENG", "ENGLISH"),
            ("OPT1", "HINDI / IT"),
            ("ECO", "ECONOMICS"),
            ("BK", "BOOK KEEPING"),
            ("OC", "ORGANISATION OF COMMERCE"),
            ("OPT2", "MATHS / SP"),
            ("EVS", "EVS"),
            ("PE", "PE"),
        ]

        # Sub-columns for numeric subjects
        sub_cols_numeric = ["UNIT I", "TERM I", "UNIT II", "INTERNAL", "ANNUAL", "TOTAL", "AVERAGE", "GRACE"]
        # Sub-columns for Grade subjects (EVS, PE) -> Just "GRADE" usually? 
        # Requirement: "PE & EVS are grade-only subjects". 
        
        current_col = 1
        
        # ROLL NO
        ws_detail.merge_cells(start_row=4, start_column=current_col, end_row=5, end_column=current_col)
        ws_detail.cell(4, current_col, "ROLL NO")
        current_col += 1
        
        # NAME
        ws_detail.merge_cells(start_row=4, start_column=current_col, end_row=5, end_column=current_col)
        ws_detail.cell(4, current_col, "STUDENT NAME")
        current_col += 1
        
        subject_start_cols = {} # code -> start_col

        for code, title in detail_cols_struct:
            start = current_col
            
            is_grade_only = code in ["EVS", "PE"]
            
            if is_grade_only:
                # Single Column for Grade
                cols_to_add = ["GRADE"]
            else:
                cols_to_add = sub_cols_numeric

            width = len(cols_to_add)
            
            ws_detail.merge_cells(start_row=4, start_column=start, end_row=4, end_column=start + width - 1)
            ws_detail.cell(4, start, title)
            
            for i, sub in enumerate(cols_to_add):
                ws_detail.cell(5, start + i, sub)
            
            subject_start_cols[code] = start
            current_col += width

        # FINAL COLUMNS
        final_headers = ["TOTAL GRACE", "TOTAL", "%", "RESULT"]
        for h in final_headers:
            ws_detail.merge_cells(start_row=4, start_column=current_col, end_row=5, end_column=current_col)
            ws_detail.cell(4, current_col, h)
            current_col += 1

        # Apply Header Styles
        for r in range(4, 6):
            for c in range(1, current_col):
                cell = ws_detail.cell(r, c)
                cell.font = header_font if r==4 else sub_header_font
                cell.fill = header_fill if r==4 else sub_header_fill
                cell.alignment = header_align
                cell.border = thin_border
        
        # --- Detail Data Rows ---
        row = 6
        for s in div_students:
            ws_detail.cell(row, 1, s.roll_no).alignment = center_align
            ws_detail.cell(row, 2, s.name).alignment = left_align
            
            res_obj = result_map.get(s.roll_no)
            
            # Helper to get numeric marks breakdown
            def fill_marks(code_key, start_col):
                # Identify actual subject code for this student
                actual_code = code_key
                if code_key == "OPT1": actual_code = s.optional_subject # e.g. HINDI
                if code_key == "OPT2": actual_code = s.optional_subject_2 # e.g. SP
                
                if not actual_code:
                    return # Empty cells

                # For Marks breakdown, use Mark table
                m = mark_map.get((s.roll_no, actual_code))
                
                # Fetch Grace for this subject from Result
                sub_grace = 0.0
                if res_obj:
                     _, sub_grace = res_obj.get_subject_data(actual_code)  # returns (avg, grace)

                vals = [
                    m.unit1 if m else 0,   # Col 0: UNIT I
                    m.term if m else 0,    # Col 1: TERM I
                    m.unit2 if m else 0,   # Col 2: UNIT II
                    m.internal if m else 0,# Col 3: INTERNAL
                    m.annual if m else 0,  # Col 4: ANNUAL
                ]
                
                # Fill Raw Marks
                for i, v in enumerate(vals):
                    c = ws_detail.cell(row, start_col + i, v)
                    c.alignment = center_align
                    c.border = thin_border

                # Col 5: TOTAL (Sum of Unit, Term, Unit2, Internal, Annual)
                start_let = get_column_letter(start_col)
                end_let = get_column_letter(start_col + 4) # Annual is 4th index (5th col)
                ws_detail.cell(row, start_col + 5, f"=SUM({start_let}{row}:{end_let}{row})").border = thin_border
                
                # Col 6: AVERAGE (DB sub_avg) - The rounded final mark out of 100
                avg_val = m.sub_avg if m else 0
                ws_detail.cell(row, start_col + 6, avg_val).border = thin_border

                # Col 7: GRACE (from Result)
                ws_detail.cell(row, start_col + 7, sub_grace).border = thin_border

            # Fill Subjects
            for code_key, start in subject_start_cols.items():
                if code_key in ["EVS", "PE"]:
                    # Grade Only
                    grade_val = ""
                    if res_obj:
                        if code_key == "EVS": grade_val = res_obj.evs_grade
                        if code_key == "PE": grade_val = res_obj.pe_grade
                    
                    c = ws_detail.cell(row, start, grade_val or "-")
                    c.alignment = center_align
                    c.border = thin_border
                else:
                    fill_marks(code_key, start)

            # Final Results
            if res_obj:
                ws_detail.cell(row, current_col-4, res_obj.total_grace).border = thin_border
                ws_detail.cell(row, current_col-3, res_obj.overall_tot).border = thin_border
                ws_detail.cell(row, current_col-2, res_obj.percentage).border = thin_border
                # RESULT Column: Use predefined overall_grade
                res_str = res_obj.overall_grade if res_obj.overall_grade else "-"
                ws_detail.cell(row, current_col-1, res_str).border = thin_border

            # Apply specific border to Name/Roll
            ws_detail.cell(row, 1).border = thin_border
            ws_detail.cell(row, 2).border = thin_border
            
            row += 1
            
        # Autosize
        autosize_columns(ws_detail, data_start_row=6)


        # ==========================================
        # 2. DIV_{X}_All (Summary Ledger)
        # ==========================================
        ws_all = wb.create_sheet(f"DIV_{div}_All")

        # Titles
        ws_all.merge_cells("A1:Q1")
        ws_all["A1"] = "SIES JUNIOR COLLEGE OF COMMERCE NERUL"
        ws_all["A1"].font = Font(bold=True, size=12)
        ws_all["A1"].alignment = center_align
        
        # Headers
        # Ref: ROLL NO, NAME, ENG, Eng Grace, Hindi/IT, Grace, ECO, Grace, BK, Grace, OC, Grace, MATH/SP, Grace, TOTAL, Grace Total, Per., Result
        # Wait, reference showed: ['ROLL NO', 'STUDENT NAME', 'ENG ', 'Eng Grace', ' Hindi', 'Hindi Grace', ...]
        
        # Note: summary_headers moved to top of function to avoid being conditionally bound
        
        for i, h in enumerate(summary_headers, 1):
            c = ws_all.cell(2, i, h)
            c.font = summary_header_font
            c.fill = summary_header_fill
            c.border = thin_border
            c.alignment = center_align

        row_all = 3
        
        for s in div_students:
            res = result_map.get(s.roll_no)
            
            # Prepare row data
            # Helper to safely get avg/grace
            def get_ag(code_slot):
                if not res: return 0.0, 0.0
                if code_slot == "ENG": return res.eng_avg, res.eng_grace
                if code_slot == "ECO": return res.eco_avg, res.eco_grace
                if code_slot == "BK": return res.bk_avg, res.bk_grace
                if code_slot == "OC": return res.oc_avg, res.oc_grace
                if code_slot == "OPT1": return res.opt1_avg, res.opt1_grace
                if code_slot == "OPT2": return res.opt2_avg, res.opt2_grace
                return 0.0, 0.0

            eng_a, eng_g = get_ag("ENG")
            opt1_a, opt1_g = get_ag("OPT1") # HINDI / IT
            eco_a, eco_g = get_ag("ECO")
            bk_a, bk_g = get_ag("BK")
            oc_a, oc_g = get_ag("OC")
            opt2_a, opt2_g = get_ag("OPT2") # MATHS / SP
            
            res_status = "-"
            if res:
                res_status = res.overall_grade if res.overall_grade else "-"
            
            data_row = [
                s.roll_no, s.name,
                eng_a, eng_g,
                opt1_a, opt1_g,
                eco_a, eco_g,
                bk_a, bk_g,
                oc_a, oc_g,
                opt2_a, opt2_g,
                res.overall_tot if res else 0,
                res.total_grace if res else 0,
                res.percentage if res else 0,
                res_status
            ]
            
            # Save for Sheet2
            if div not in all_summary_rows_by_div:
                all_summary_rows_by_div[div] = []
            all_summary_rows_by_div[div].append(data_row)
            
            # Write to DIV_All
            for i, val in enumerate(data_row, 1):
                c = ws_all.cell(row_all, i, val)
                c.border = thin_border
                c.alignment = center_align if i > 2 else left_align
            
            row_all += 1

        autosize_columns(ws_all, data_start_row=3)

        # ==========================================
        # 3. NAME {X} (Name List)
        # ==========================================
        ws_name = wb.create_sheet(f"NAME {div}")
        ws_name["A1"] = "ROLL NO"
        ws_name["B1"] = "STUDENT NAME"
        
        ws_name["A1"].font = sub_header_font
        ws_name["B1"].font = sub_header_font
        
        r_name = 2
        for s in div_students:
            ws_name.cell(r_name, 1, s.roll_no)
            ws_name.cell(r_name, 2, s.name)
            r_name += 1
            
        ws_name.column_dimensions["A"].width = 15
        ws_name.column_dimensions["B"].width = 30


    # ==========================================
    # 4. Failed Students (Consolidated Batch Summary)
    # ==========================================
    ws_sheet2 = wb.create_sheet("Fail List")
    
    # Title
    ws_sheet2.merge_cells("A1:Q1")
    ws_sheet2["A1"] = "SIES JUNIOR COLLEGE OF COMMERCE, NERUL"
    ws_sheet2["A1"].font = Font(bold=True, size=14)
    ws_sheet2["A1"].alignment = center_align

    row_s2 = 3
    
    for div in divisions:
        div_rows = all_summary_rows_by_div.get(div, [])
        failed_rows = []
        for r in div_rows:
            res_val = str(r[-1]).strip().upper() if r[-1] else ""
            if "FAIL" in res_val:
                failed_rows.append(r)
        
        if not failed_rows:
            continue

        # Division Header
        ws_sheet2.merge_cells(start_row=row_s2, start_column=1, end_row=row_s2, end_column=len(summary_headers))
        ws_sheet2.cell(row_s2, 1, f"DIVISION {div}")
        ws_sheet2.cell(row_s2, 1).font = Font(bold=True, size=12)
        ws_sheet2.cell(row_s2, 1).alignment = center_align
        ws_sheet2.cell(row_s2, 1).fill = sub_header_fill
        row_s2 += 1
        
        # Table Headers
        for i, h in enumerate(summary_headers, 1):
            c = ws_sheet2.cell(row_s2, i, h)
            c.font = summary_header_font
            c.fill = summary_header_fill
            c.border = thin_border
            c.alignment = center_align
        row_s2 += 1
        
        # Data
        for data_row in failed_rows:
            for i, val in enumerate(data_row, 1):
                c = ws_sheet2.cell(row_s2, i, val)
                c.border = thin_border
                c.alignment = center_align if i > 2 else left_align
            row_s2 += 1
            
        # Spacing
        row_s2 += 2

    autosize_columns(ws_sheet2, data_start_row=3)

    return wb
